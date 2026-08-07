"""
Excel workbook generator (openpyxl).

Produces the SAME styled workbook the customer's existing inventory script emits — a
navy/Databricks-red Summary sheet plus one sheet per fine-grained asset type, with the same
columns and cell formatting — so the output matches what the customer is already used to.
The styling + column layout are a faithful port of that script (`_render_excel`); asset
metadata (labels, columns, formatters) is shared with the HTML via `inventory_view.py`.

Additions over the reference (deliberate):
  • UC / MLflow sheets are omitted (out of scope for this migration utility).
  • A "Migration Plan" checklist sheet is appended (this utility's convention).

IMPORTANT: openpyxl needs a seekable local disk — do NOT write straight to a FUSE /Volumes
path (it corrupts). Callers render to a local /tmp path via ArtifactWriter and byte-copy.
"""
from __future__ import annotations

import re
from datetime import datetime

from src.reports.inventory_view import (
    _COLUMNS,
    _LABELS,
    _SUMMARY_CARD_KEYS,
    _deep_get,
    _resolve_items,
    adapt,
    build_counts,
)

_EXCEL_HEADER_BG = "1E3A5F"   # deep navy
_EXCEL_SECTION_BG = "334155"  # slate
_EXCEL_DB_RED = "FF3621"      # Databricks brand red
_EXCEL_ALT_ROW = "F1F5F9"     # very light gray


def _cell_text(value, fmt: str):
    """Plain-text / native value for an Excel cell — ported from the reference script."""
    # Tri-state columns are checked BEFORE the empty guard: for them `None` is meaningful
    # ("could not check"), not missing, so it must not fall through to a blank cell.
    if fmt == "badge_bool_unknown":
        if value is True or str(value).lower() in ("true", "1", "yes"):
            return "Yes"
        if value is False or str(value).lower() in ("false", "0", "no"):
            return "No"
        return "Could not check"
    if value is None or value == "":
        return ""
    if fmt in ("plain", "mono", "path", "trunc"):
        s = str(value)
        return s[:200] if fmt == "trunc" else s
    if fmt == "short_mono":
        return str(value)[:8]
    if fmt == "badge_bool":
        return "Yes" if (value is True or str(value).lower() in ("true", "1", "yes")) else "No"
    if fmt in ("badge_state", "badge_type", "badge_lang", "badge_managed"):
        return str(value)
    if fmt == "cls_managed":
        from src.reports.inventory_view import managed_by_label
        return managed_by_label(value)
    if fmt == "count":
        return len(value) if isinstance(value, list) else 0
    if fmt == "first_email":
        if isinstance(value, list) and value:
            return value[0].get("value", "")
        return ""
    if fmt == "list_vals":
        if isinstance(value, list) and value:
            return ", ".join(str(v.get("value", v)) for v in value)
        return ""
    if fmt == "schedule":
        if isinstance(value, dict):
            cron = value.get("quartz_cron_expression", "")
            tz = value.get("timezone_id", "")
            return f"{cron}  ({tz})" if cron else "Manual"
        return "Manual"
    if fmt == "epoch_ms":
        try:
            return datetime.fromtimestamp(int(value) / 1000).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(value)
    if fmt == "iso_ts":
        try:
            return str(value)[:16].replace("T", " ")
        except Exception:
            return str(value)
    if fmt == "url_link":
        return str(value)
    if fmt == "kv_dns":
        return value.get("dns_name", "") if isinstance(value, dict) else ""
    return str(value)


def generate_excel(objects_by_type: dict, counts: dict = None, local_path: str = "",
                   config=None) -> str:
    """Write the reference-style workbook to `local_path` (a real seekable path); return it.

    `objects_by_type` is our collectors' output; it is adapted to the reference renderer's
    per-asset shape internally. `counts` is kept for signature compatibility but recomputed.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workspace_url = getattr(getattr(config, "ctx", None), "workspace_url", "") if config else ""
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    data = adapt(objects_by_type or {})
    counts = build_counts(data)

    def _fill(c):
        return PatternFill("solid", fgColor=c)

    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    last_col = get_column_letter(3)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "Databricks Workspace Inventory"
    c.font = _font(bold=True, color="FFFFFF", size=16)
    c.fill = _fill(_EXCEL_DB_RED)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Workspace: {workspace_url}   |   Generated: {generated_at}"
    c.font = _font(italic=True, color="475569", size=9)
    c.fill = _fill("FFF5F5")
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for col_idx, header in enumerate(["Component", "Count", "Excel Sheet"], 1):
        c = ws.cell(row=4, column=col_idx, value=header)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(_EXCEL_HEADER_BG)
        c.alignment = _align("center")
        c.border = box
    ws.row_dimensions[4].height = 22

    total = 0
    for row_idx, key in enumerate(_SUMMARY_CARD_KEYS, 5):
        count = counts.get(key, 0)
        total += count
        label = _LABELS.get(key, key.replace("_", " ").title())
        sheet_name = re.sub(r'[\\/?*\[\]:]', '-', label)[:31]
        bg = _EXCEL_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, val in enumerate([label, count, sheet_name], 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = _fill(bg)
            c.border = box
            c.font = _font(bold=(col_idx == 2), size=10)
            c.alignment = _align("center" if col_idx == 2 else "left")

    total_row = len(_SUMMARY_CARD_KEYS) + 5
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = _font(bold=True, size=10, color="FFFFFF")
    c.fill = _fill(_EXCEL_DB_RED)
    c.border = box
    c.alignment = _align("center")
    c2 = ws.cell(row=total_row, column=2, value=total)
    c2.font = _font(bold=True, size=11, color="FFFFFF")
    c2.fill = _fill(_EXCEL_DB_RED)
    c2.border = box
    c2.alignment = _align("center")
    c3 = ws.cell(row=total_row, column=3, value="")
    c3.fill = _fill(_EXCEL_DB_RED)
    c3.border = box
    ws.row_dimensions[total_row].height = 22

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A5"

    # ── Component sheets ─────────────────────────────────────────────────
    for key in _SUMMARY_CARD_KEYS:
        items = _resolve_items(data, key)
        cols = _COLUMNS.get(key, [("path", "Path", "plain")])
        label = _LABELS.get(key, key.replace("_", " ").title())
        sheet_name = re.sub(r'[\\/?*\[\]:]', '-', label)[:31]
        n_cols = len(cols)

        ws2 = wb.create_sheet(title=sheet_name)
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        c = ws2["A1"]
        c.value = f"{label}  —  {len(items):,} items"
        c.font = _font(bold=True, color="FFFFFF", size=12)
        c.fill = _fill(_EXCEL_HEADER_BG)
        c.alignment = _align("left")
        ws2.row_dimensions[1].height = 26

        ws2.merge_cells(f"A2:{get_column_letter(n_cols)}2")
        c = ws2["A2"]
        c.value = f"Workspace: {workspace_url}   |   Generated: {generated_at}"
        c.font = _font(italic=True, color="64748B", size=9)
        c.fill = _fill("F8FAFC")
        c.alignment = _align("left")
        ws2.row_dimensions[2].height = 14
        ws2.row_dimensions[3].height = 4

        for col_idx, (_, col_label, _fmt) in enumerate(cols, 1):
            c = ws2.cell(row=4, column=col_idx, value=col_label)
            c.font = _font(bold=True, color="FFFFFF", size=10)
            c.fill = _fill(_EXCEL_SECTION_BG)
            c.alignment = _align("center")
            c.border = box
        ws2.row_dimensions[4].height = 20
        ws2.freeze_panes = "A5"

        col_widths = [len(col_label) for (_, col_label, _fmt) in cols]
        for row_idx, item in enumerate(items, 5):
            bg = _EXCEL_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, (key_path, _lbl, fmt) in enumerate(cols, 1):
                text = _cell_text(_deep_get(item, key_path), fmt)
                c = ws2.cell(row=row_idx, column=col_idx, value=text)
                c.fill = _fill(bg)
                c.font = _font(size=10)
                c.alignment = _align("left", "top")
                c.border = box
                if row_idx <= 104:
                    col_widths[col_idx - 1] = max(
                        col_widths[col_idx - 1], min(len(str(text)) if text else 0, 60))

        for col_idx, width in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 68)

    # ── Migration Plan checklist sheet (this utility's addition) ──────────
    mp = wb.create_sheet(title="Migration Plan")
    mp.sheet_view.showGridLines = False
    for col_idx, header in enumerate(["Component", "Natural key", "Status (fill in)", "Notes"], 1):
        c = mp.cell(row=1, column=col_idx, value=header)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill(_EXCEL_HEADER_BG)
        c.alignment = _align("center")
        c.border = box
    mp.freeze_panes = "A2"
    row = 2
    for otype, objs in (objects_by_type or {}).items():
        for o in objs or []:
            if not isinstance(o, dict):
                continue
            nk = (o.get("natural_key") or o.get("_natural_key")
                  or o.get("name") or o.get("path") or o.get("displayName") or "")
            mp.cell(row=row, column=1, value=otype)
            mp.cell(row=row, column=2, value=str(nk))
            row += 1
    mp.column_dimensions["A"].width = 22
    mp.column_dimensions["B"].width = 48
    mp.column_dimensions["C"].width = 18
    mp.column_dimensions["D"].width = 40

    wb.save(local_path)
    return local_path
