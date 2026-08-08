"""
export_excel — `export_status.xlsx`: the inventory workbook + an Export Status column (Plan 2 §6a).

The operator's post-export checkpoint. Base = the SAME per-asset sheets/columns/styling as the
inventory workbook (reuses `inventory_view` metadata + `excel_generator` styling), with TWO columns
added to every per-asset sheet, both joined from `export_index.json` on `(asset_type, natural_key)`:

  • **Export Status** — did EXPORT capture this unit? (Success / Skipped (DAB) / Manual / …)
  • **Import Action** — what will the TARGET side DO with it? (CREATE / ASSIGN / DAB REDEPLOY / …)

Two columns because one can't carry both meanings: a bundle-owned job is "Skipped (DAB)" on export
yet still lands on target — via the customer's bundle redeploy, which only the action column says.
Originally the action column existed on the identity sheets alone, which left every other tab's
intent to be guessed from its status.

Plus a leading **Export Summary** sheet: per-asset_type status counts, an Import Action roll-up
(how much the tool does vs. the bundle vs. a human), a failures table (red, reasons first), and a
separate "Oversize — manual copy needed" table.

Every inventoried row gets a status (no blank cells) — the human-readable "exported True/False"
tie-back. Rendered to a local /tmp path then byte-copied to the Volume (openpyxl-on-FUSE gotcha).
"""
from __future__ import annotations

import json
import re
from datetime import datetime

from src.exporters.asset_export import (
    DAB_CONTENT_NOTE,
    IMPORT_ACTIONS,
    _ACTION_DAB,
    is_dab_content_path,
)
from src.reports.inventory_view import (
    _COLUMNS,
    _LABELS,
    _SUMMARY_CARD_KEYS,
    _deep_get,
    _resolve_items,
    adapt,
)

# Inventory card key → export asset_type (single). Multi-type cards handled in _row_asset_type.
_CARD_ASSET_TYPE = {
    "users": "user", "groups": "group", "service_principals": "service_principal",
    "notebooks": "notebook", "workspace_files": "workspace_file", "sql_queries": "legacy_query",
    "jobs": "job", "clusters": "cluster", "instance_pools": "instance_pool",
    "cluster_policies": "cluster_policy", "cluster_libraries": "cluster_library",
    "global_init_scripts": "global_init_script", "sql_warehouses": "sql_warehouse",
    "sql_dashboards": "legacy_dashboard", "dlt_pipelines": "dlt_pipeline",
    "lakeview_dashboards": "lakeview_dashboard", "genie_spaces": "genie_space",
    "serving_endpoints": "serving_endpoint", "secret_scopes": "secret_scope",
    "repos": "repo", "workspace_conf": "workspace_conf", "apps": "app",
    "lakebase_projects": "lakebase_project",
}

# Per-card ordered natural-key candidate fields (dotted → _deep_get) on the adapted row.
_CARD_NK_FIELDS = {
    "users": ["userName"], "groups": ["displayName"], "service_principals": ["applicationId"],
    "notebooks": ["path"], "workspace_files": ["path"], "repos": ["path"],
    "sql_queries": ["display_name", "name"], "jobs": ["settings.name", "name"],
    "clusters": ["cluster_name"], "instance_pools": ["instance_pool_name"],
    "cluster_policies": ["name"], "global_init_scripts": ["name"],
    "sql_warehouses": ["name"], "sql_dashboards": ["name", "display_name"],
    "dlt_pipelines": ["name"], "lakeview_dashboards": ["display_name"],
    "genie_spaces": ["title"], "serving_endpoints": ["name"], "secret_scopes": ["name"],
    "workspace_conf": ["key"], "apps": ["name"], "lakebase_projects": ["name"],
}

# Export status → (display label, fill hex). Mirrors §6a.
_STATUS_STYLE = {
    "success": ("Success", "D1FAE5"),
    "failure": ("Failure", "FEE2E2"),
    "skip": ("Skip", "E5E7EB"),
    "manual": ("Manual", "FEF3C7"),
    # Bundle-owned: the asset IS in the ledger, but its create payload is intentionally not
    # captured (the customer's bundle redeploy owns the definition) — so it reads as a
    # deliberate skip, with the Import Action column naming who recreates it.
    "dab": ("Skipped (DAB)", "DBEAFE"),
    "covered": ("Covered (native)", "CFFAFE"),
    "skipped_oversize": ("Skipped (oversize) ⚠", "FDE68A"),
    "incomplete": ("Incomplete", "FED7AA"),
    "": ("—", "FFFFFF"),
}


def _row_asset_type(card_key: str, row: dict) -> str:
    """Resolve a row's export asset_type; the sql_alerts card mixes legacy vs V2."""
    if card_key == "sql_alerts":
        return "alert_v2" if str(row.get("_alert_kind")) == "Alerts V2" else "legacy_alert"
    return _CARD_ASSET_TYPE.get(card_key, "")


def _row_natural_key(card_key: str, row: dict) -> str:
    fields = _CARD_NK_FIELDS.get(card_key)
    if card_key == "sql_alerts":
        fields = ["display_name", "name"]
    for f in fields or []:
        v = _deep_get(row, f)
        if v not in (None, ""):
            return str(v)
    return ""


def _status_lookup(index: dict) -> dict:
    """(asset_type, natural_key) → (status, note) from export_index.json."""
    out = {}
    for r in index.get("units", []) or []:
        out[(r.get("asset_type"), r.get("natural_key"))] = (r.get("export_status", ""),
                                                            r.get("note", ""))
    return out


# EVERY per-asset sheet gets an Import Action column: `export_status` only says "we captured it",
# which must not be read as "the tool will create it on target" — and for bundle-owned assets
# (which now report Success like everything else) the action column is the ONLY place the
# "your bundle redeploys this, import skips it" instruction appears.
_IMPORT_ACTION_LABEL = {
    "create": "CREATE on target",
    "create_and_upload": "CREATE + UPLOAD content",
    "assign_on_target": "ASSIGN (must pre-exist in account)",
    "adopt_or_assign": "AUTO — adopt/assign (no action needed)",
    "add_members": "ADD MEMBERS (group exists)",
    "dab_redeploy": "DAB REDEPLOY (import skips)",
    "via_native_asset": "NONE — via native asset",
    "install": "INSTALL on target",
    "set_conf": "SET on target",
    "apply_acl": "APPLY ACL on target",
    "manual": "MANUAL on target",
    "review_required": "REVIEW REQUIRED",
    "skip_generated": "SKIP — Databricks-generated (nothing to do)",
    "none": "NOT EXPORTED (nothing to import)",
    "": "—",
}
_IMPORT_ACTION_FILL = {
    "create": "DCFCE7",              # green — the utility does it
    "create_and_upload": "DCFCE7",   # green — the utility does it (create + bytes)
    "add_members": "DCFCE7",         # green — the utility does it (PATCH members)
    "install": "DCFCE7",             # green — the utility does it (attach to existing cluster)
    "set_conf": "DCFCE7",            # green — the utility does it (conf API)
    "apply_acl": "DCFCE7",           # green — the utility does it (permissions API)
    "adopt_or_assign": "DCFCE7",     # green — the utility does it (POST adopts/assigns; appId kept)
    "assign_on_target": "DBEAFE",    # blue  — account/IT prerequisite (account GROUPS only)
    "dab_redeploy": "DBEAFE",        # blue  — the customer's bundle pipeline owns it
    "via_native_asset": "CFFAFE",    # cyan  — happens as a side effect, no separate action
    "manual": "FEF3C7",              # amber — a human must do it on target
    "review_required": "FEF3C7",     # amber — human must confirm
    "skip_generated": "E5E7EB",      # grey  — platform artifact, nothing to do
    "none": "E5E7EB",                # grey  — nothing to import
    "": "F1F5F9",
}

# Actions whose cell should also show the unit's note: the note is where the actual caveat lives
# ("UC tables must pre-exist", "copy the DBFS jar by hand", "client secret not exportable").
_ACTION_SHOW_NOTE = {"manual", "dab_redeploy", "review_required", "assign_on_target",
                     "adopt_or_assign", "create", "install"}

# Fail loudly if a new action is added to the producer without a label here — otherwise it would
# render as "—", which is the blank-cell failure mode the customer explicitly rejected.
_missing_labels = IMPORT_ACTIONS - set(_IMPORT_ACTION_LABEL)
assert not _missing_labels, f"import_action(s) with no Excel label: {sorted(_missing_labels)}"


def _import_action_lookup(index: dict) -> dict:
    """(asset_type, natural_key) → import_action, for every per-asset sheet."""
    out = {}
    for r in index.get("units", []) or []:
        if r.get("import_action"):
            out[(r.get("asset_type"), r.get("natural_key"))] = r["import_action"]
    return out


# NOTE: the card→asset_type and card→natural-key mappings already exist above as
# _CARD_ASSET_TYPE / _CARD_NK_FIELDS and are reused here — do NOT redefine them (an earlier
# identity-only redefinition shadowed the full map and silently blanked the Export Status column
# for every non-identity card).


def _cluster_library_status(row: dict, index_units: list) -> tuple[str, str, str]:
    """cluster_library natural_key embeds a JSON blob — join by cluster_id + library_name substring."""
    cid = str(row.get("cluster_id") or "")
    lib_name = str(row.get("library_name") or "")
    for u in index_units:
        if u.get("asset_type") != "cluster_library":
            continue
        nk = str(u.get("natural_key") or "")
        if nk.startswith(f"{cid}:") and (not lib_name or lib_name in nk):
            return u.get("export_status", ""), u.get("note", ""), u.get("import_action", "")
    return "", "", ""


def generate_export_excel(objects_by_type: dict, index: dict, local_path: str,
                          config=None) -> str:
    """Write the export-status workbook to `local_path` (seekable); return it."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workspace_url = getattr(getattr(config, "ctx", None), "workspace_url", "") if config else ""
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    data = adapt(objects_by_type or {})
    status_by_key = _status_lookup(index)
    action_by_key = _import_action_lookup(index)
    index_units = index.get("units", []) or []

    def _fill(c):
        return PatternFill("solid", fgColor=c)

    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Export Summary sheet ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Export Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Workspace Export — Status Summary"
    c.font = _font(bold=True, color="FFFFFF", size=16)
    c.fill = _fill("FF3621")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (f"Workspace: {workspace_url}   |   Run: {index.get('run_id','')}   |   "
               f"Generated: {generated_at}")
    c.font = _font(italic=True, color="475569", size=9)
    c.alignment = _align("center")

    status_cols = ["success", "failure", "skip", "manual", "dab", "covered",
                   "skipped_oversize", "incomplete"]
    # Explicit short headers — deriving them from _STATUS_STYLE's first word made "Skipped (DAB)"
    # and "Skipped (oversize)" both render as "Skipped", two indistinguishable columns.
    _SHORT = {"success": "Success", "failure": "Failure", "skip": "Skip", "manual": "Manual",
              "dab": "DAB", "covered": "Covered", "skipped_oversize": "Oversize",
              "incomplete": "Incomplete"}
    headers = ["Asset Type", "Total"] + [_SHORT[s] for s in status_cols]
    for col_idx, h in enumerate(headers, 1):
        cc = ws.cell(row=4, column=col_idx, value=h)
        cc.font = _font(bold=True, color="FFFFFF", size=10)
        cc.fill = _fill("1E3A5F")
        cc.alignment = _align("center")
        cc.border = box
    counts = index.get("counts", {}) or {}
    row_idx = 5
    for at in sorted(counts):
        bucket = counts[at]
        vals = [at, bucket.get("total", 0)] + [bucket.get(s, 0) for s in status_cols]
        for col_idx, v in enumerate(vals, 1):
            cc = ws.cell(row=row_idx, column=col_idx, value=v)
            cc.font = _font(size=10, bold=(col_idx == 2))
            cc.border = box
            cc.alignment = _align("left" if col_idx == 1 else "center")
        row_idx += 1
    for i, w in enumerate([30, 8, 9, 9, 7, 7, 7, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Failures table.
    failures = [r for r in index_units if r.get("export_status") == "failure"]
    fr = row_idx + 2
    ws.cell(row=fr, column=1, value=f"Failures ({len(failures)})").font = _font(bold=True, color="B91C1C", size=12)
    fr += 1
    for col_idx, h in enumerate(["Asset Type", "Natural Key", "Reason"], 1):
        cc = ws.cell(row=fr, column=col_idx, value=h)
        cc.font = _font(bold=True, color="FFFFFF")
        cc.fill = _fill("B91C1C")
        cc.border = box
    for r in failures:
        fr += 1
        for col_idx, v in enumerate([r.get("asset_type"), r.get("natural_key"), r.get("note")], 1):
            cc = ws.cell(row=fr, column=col_idx, value=v)
            cc.fill = _fill("FEE2E2")
            cc.border = box
            cc.font = _font(size=10)

    # Import Action roll-up: "how much does the tool do vs. the bundle vs. a human?" — the one
    # question the per-type status table can't answer, since one type can span several actions.
    action_counts = index.get("action_counts") or {}
    if not action_counts:
        # Older index (written before action_counts existed) → derive it from the units.
        action_counts = {}
        for r in index_units:
            a = r.get("import_action") or ""
            action_counts[a] = action_counts.get(a, 0) + 1
    ar = fr + 2
    ws.cell(row=ar, column=1, value="Import Actions (what the target side will do)").font = \
        _font(bold=True, color="1E3A5F", size=12)
    ar += 1
    for col_idx, h in enumerate(["Import Action", "Count"], 1):
        cc = ws.cell(row=ar, column=col_idx, value=h)
        cc.font = _font(bold=True, color="FFFFFF")
        cc.fill = _fill("334155")
        cc.border = box
    for act in sorted(action_counts, key=lambda a: (-action_counts[a], a)):
        ar += 1
        cc = ws.cell(row=ar, column=1, value=_IMPORT_ACTION_LABEL.get(act, act or "—"))
        cc.fill = _fill(_IMPORT_ACTION_FILL.get(act, _IMPORT_ACTION_FILL[""]))
        cc.border = box
        cc.font = _font(size=10)
        cc = ws.cell(row=ar, column=2, value=action_counts[act])
        cc.border = box
        cc.font = _font(size=10, bold=True)
        cc.alignment = _align("center")
    fr = ar

    # Oversize table.
    oversize = [r for r in index_units if r.get("export_status") == "skipped_oversize"]
    orow = fr + 2
    ws.cell(row=orow, column=1,
            value=f"Oversize — manual copy needed ({len(oversize)})").font = _font(bold=True, color="B45309", size=12)
    orow += 1
    for col_idx, h in enumerate(["Asset Type", "Path", "Note"], 1):
        cc = ws.cell(row=orow, column=col_idx, value=h)
        cc.font = _font(bold=True, color="FFFFFF")
        cc.fill = _fill("B45309")
        cc.border = box
    for r in oversize:
        orow += 1
        for col_idx, v in enumerate([r.get("asset_type"), r.get("natural_key"), r.get("note")], 1):
            cc = ws.cell(row=orow, column=col_idx, value=v)
            cc.fill = _fill("FDE68A")
            cc.border = box
            cc.font = _font(size=10)

    # ── Per-asset sheets (inventory columns + Export Status) ──────────────
    from src.exporters.excel_generator import _cell_text
    for key in _SUMMARY_CARD_KEYS:
        items = _resolve_items(data, key)
        cols = _COLUMNS.get(key, [("path", "Path", "plain")])
        label = _LABELS.get(key, key.replace("_", " ").title())
        sheet_name = re.sub(r'[\\/?*\[\]:]', '-', label)[:31]
        n_cols = len(cols) + 2   # + Export Status + Import Action (every sheet)

        ws2 = wb.create_sheet(title=sheet_name)
        ws2.sheet_view.showGridLines = False
        ws2.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        c = ws2["A1"]
        c.value = f"{label}  —  {len(items):,} items"
        c.font = _font(bold=True, color="FFFFFF", size=12)
        c.fill = _fill("1E3A5F")
        c.alignment = _align("left")
        ws2.row_dimensions[1].height = 24

        header_cells = [cl for (_, cl, _f) in cols] + ["Export Status", "Import Action"]
        for col_idx, col_label in enumerate(header_cells, 1):
            cc = ws2.cell(row=3, column=col_idx, value=col_label)
            cc.font = _font(bold=True, color="FFFFFF", size=10)
            cc.fill = _fill("334155")
            cc.alignment = _align("center")
            cc.border = box
        ws2.freeze_panes = "A4"

        col_widths = [len(cl) for cl in header_cells]
        for row_num, item in enumerate(items, 4):
            for col_idx, (key_path, _lbl, fmt) in enumerate(cols, 1):
                text = _cell_text(_deep_get(item, key_path), fmt)
                cc = ws2.cell(row=row_num, column=col_idx, value=text)
                cc.font = _font(size=10)
                cc.alignment = _align("left", "top")
                cc.border = box
                if row_num <= 103:
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1],
                                                  min(len(str(text)) if text else 0, 60))
            # Export Status cell ("did export capture this?").
            status, note, action = _resolve_status(key, item, status_by_key, index_units,
                                                   action_by_key)
            slabel, sfill = _STATUS_STYLE.get(status, _STATUS_STYLE[""])
            scell = ws2.cell(row=row_num, column=n_cols - 1,
                             value=slabel + (f" — {note}" if note and status in ("failure",) else ""))
            scell.fill = _fill(sfill)
            scell.border = box
            scell.font = _font(size=10, bold=(status == "failure"))
            # Import Action cell ("what will the TARGET side do with it?") — on every sheet, and
            # the only place a DAB row says who recreates it. The note rides along for the
            # actions where the caveat is the whole point (UC prereqs, DBFS artifacts, secrets).
            acell = ws2.cell(
                row=row_num, column=n_cols,
                value=_IMPORT_ACTION_LABEL.get(action, action or "—")
                + (f" — {note}" if note and action in _ACTION_SHOW_NOTE else ""))
            acell.fill = _fill(_IMPORT_ACTION_FILL.get(action, _IMPORT_ACTION_FILL[""]))
            acell.border = box
            acell.alignment = _align("left", "top", wrap=True)
            acell.font = _font(size=10, bold=(action in ("review_required", "manual")))
        for col_idx, w in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(w + 4, 68)
        ws2.column_dimensions[get_column_letter(n_cols - 1)].width = 22
        ws2.column_dimensions[get_column_letter(n_cols)].width = 46

    wb.save(local_path)
    return local_path


# ACL-sheet `object_type` (lower-cased from the workspace API's DIRECTORY/NOTEBOOK/FILE) → the
# export asset_type `is_dab_content_path` expects. Only workspace content can sit inside a bundle
# root; a job or warehouse is bundle-owned via `dab_registry` and its own row already says so.
_ACL_WS_OBJECT_TYPES = {"directory": "directory", "notebook": "notebook", "file": "workspace_file"}


def _is_dab_acl_row(row: dict) -> bool:
    """Whether an Object-Permissions row grants on workspace content inside a bundle root."""
    asset_type = _ACL_WS_OBJECT_TYPES.get(str(row.get("object_type") or "").lower())
    if not asset_type:
        return False
    return is_dab_content_path(asset_type, row.get("object_key"))


def _resolve_status(card_key: str, row: dict, status_by_key: dict, index_units: list,
                    action_by_key: dict):
    """(status, note, import_action) for one inventory row, joined to the export index."""
    if card_key == "object_permissions":
        # Every grant is captured wholesale in acls.json, so the STATUS is uniformly success —
        # export really did capture it. The ACTION is not uniform: the importer replays ACLs only
        # for objects it created, and it deliberately creates nothing under a bundle root, so a
        # grant on bundle content reads DAB REDEPLOY like the bundle-owned jobs do. Claiming
        # "APPLY ACL on target" there promised an action import will never take.
        if _is_dab_acl_row(row):
            return "success", DAB_CONTENT_NOTE, _ACTION_DAB
        return "success", "", "apply_acl"
    if card_key == "cluster_libraries":
        return _cluster_library_status(row, index_units)
    asset_type = _row_asset_type(card_key, row)
    nk = _row_natural_key(card_key, row)
    if not asset_type:
        return "", "", ""
    status, note = status_by_key.get((asset_type, nk), ("", ""))
    return status, note, action_by_key.get((asset_type, nk), "")
