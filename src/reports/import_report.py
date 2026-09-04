"""
import_report — everything `04_Import` outputs (Plan 3 §1a, D16; slimmed in PLAN 7 §B2).

Import owns its OWN complete, customer-readable output set — it is not "test now, report later".
What Plan 4 adds is only the cross-stage inventoried→exported→imported join, which needs nothing
new from here because `import_results.json` is deliberately written in the shape Plan 4 joins on:
one row per unit, keyed `(asset_type, natural_key)`.

The output set was slimmed (PLAN 7 §B2) to the reporting chain that actually matters — inventory
(the complete list) → export_status (per-asset export status) → import_status (per-asset import
status) — plus the machine-readable join surface. So:

  • `misc/import_results.json`   — machine-readable per-unit outcome (the join surface Plan 4 reads)
  • `reports/import_status.xlsx` — the operator's artifact: Summary + a sheet per asset type +
                                   an "ACL Parity" sheet (folded in from the ACL phase, D-1), with
                                   Import Status / Action Taken / Target Id / Note columns.
                                   On a DRY RUN it is written as `import_status_dry_run.xlsx` (A1) so
                                   a live run's report never overwrites the rehearsal's.
  • `reports/manual_actions_import.md` — everything a human must still do, with reasons (a SEPARATE
                                   file from export's runbook, which still applies)

REMOVED in PLAN 7 §B2: `import_results.html` (the xlsx carries every unit + status) and the
standalone `acl_parity_report.{json,html}` (folded into the workbook). `html_generator` stays in
the tree but is no longer called from here.

The .xlsx is rendered to /tmp then byte-copied to the Volume: openpyxl needs seeks, and writing
straight to a FUSE `/Volumes` path corrupts the file (memory `uc-volume-file-io-limits`).
"""
from __future__ import annotations

from datetime import datetime

from src.exporters import bundle_paths as BP
from src.utils.helpers import now_iso, safe_str
from src.utils.logger import get_logger

_LOG = get_logger("import_report")

# import_status → (label, colour). Failures red, degraded amber, deferred grey — so the report is
# scannable and `skipped_no_object` never reads as an error (it usually isn't one).
_STATUS_STYLE = {
    "created": ("Created", "D1FAE5"),
    "created_with_warning": ("Created (warning)", "FDE68A"),
    "updated": ("Updated", "DBEAFE"),
    "adopted": ("Adopted (pre-existing)", "CFFAFE"),
    "skipped": ("Skipped (unchanged)", "E5E7EB"),
    "failed": ("FAILED", "FEE2E2"),
    "manual": ("Manual step", "FEF3C7"),
    "not_selected": ("Deferred (not selected)", "F1F5F9"),
    "skipped_no_object": ("Skipped (no target object)", "EDE9FE"),
    "deleted_in_source": ("Deleted in source", "FFE4E6"),
    "": ("—", "FFFFFF"),
}

_SUMMARY_ORDER = ("created", "updated", "adopted", "skipped", "created_with_warning", "manual",
                  "not_selected", "skipped_no_object", "failed")

# Bug 6 — one-time static disclaimer (xlsx summary footer + runbook). Account-level access-control
# is invisible to a workspace-scoped tool, so a change to it reads as `unchanged` here; this stops
# that being mistaken for a defect. See PLAN 8 "Scope clarification / Bug 6".
_ACCOUNT_ACL_DISCLAIMER = (
    "Account-level access-control — group Manager, service-principal Can use / Can manage "
    "(account rule-sets) — is managed at the account and is NOT tracked by this workspace-scoped "
    "tool. Changes to it appear here as 'unchanged'. In the same account it is already in effect on "
    "the target; across accounts it is provisioned during account setup / Entra→SCIM.")

# Fine-grained importer `asset_type` → the inventory/export CARD KEY it belongs to, so the import
# workbook lays out ONE SHEET PER ASSET TYPE exactly like inventory.xlsx / export_status.xlsx
# (IMP-1: the three stages must be structurally identical, not one-sheet-per-family here and
# one-per-type there). Any asset_type absent from this map falls back to its own name as a tab, so
# a newly-added importer type is never silently dropped.
_ASSET_TYPE_TO_CARD = {
    "user": "users", "service_principal": "service_principals",
    # PLAN 8 Bug 3: the standing OAuth-secret manual task sits beside its SP in the SPs tab.
    "service_principal_secret": "service_principals",
    "group": "groups", "group_membership": "groups",
    "notebook": "notebooks", "workspace_file": "workspace_files",
    "directory": "notebooks", "repo": "repos",
    "job": "jobs", "cluster": "clusters", "instance_pool": "instance_pools",
    "cluster_policy": "cluster_policies", "cluster_library": "cluster_libraries",
    "global_init_script": "global_init_scripts",
    "sql_warehouse": "sql_warehouses", "legacy_query": "sql_queries",
    "legacy_alert": "sql_alerts", "alert_v2": "sql_alerts",
    "legacy_dashboard": "sql_dashboards",
    "dlt_pipeline": "dlt_pipelines", "lakeview_dashboard": "lakeview_dashboards",
    "genie_space": "genie_spaces", "serving_endpoint": "serving_endpoints",
    "secret_scope": "secret_scopes", "secret_value": "secret_scopes",
    "workspace_conf": "workspace_conf", "acl": "object_permissions",
}


def _card_for_asset_type(asset_type: str) -> str:
    return _ASSET_TYPE_TO_CARD.get(safe_str(asset_type), safe_str(asset_type) or "other")


def _all_rows(results: list) -> list[dict]:
    """Every per-unit row across all phases, failures first then by (asset_type, natural_key)."""
    rows: list[dict] = []
    for res in results:
        rows.extend(res.units)
    rows.sort(key=lambda r: (safe_str(r.get("import_status")) != "failed",
                             safe_str(r.get("asset_type")), safe_str(r.get("natural_key"))))
    return rows


def write_import_reports(aw, config, summary: dict, results: list, context: dict) -> dict:
    """Write every import artifact. Never raises — a reporting failure must not fail the run.

    Report FILENAMES vary so a re-run against the SAME run dir never silently clobbers the prior
    report (each import re-uses the same bundle/run_id):
      • dry run           → `import_status_dry_run.xlsx` (A1); results json + runbook stay canonical.
      • live + retry_mode  → the WHOLE set is tagged `_retry_<timestamp>` and the canonical full-run
        (≠ off)             files (`import_status.xlsx`, `import_results.json`,
                            `manual_actions_import.md`) are LEFT UNTOUCHED — the last full run's
                            report survives, and every retry keeps its own timestamped set.
      • live + retry off  → the canonical names (the full-run report).
    The returned dict carries the ACTUAL paths written, so the notebook reads back the right files.
    """
    rows = _all_rows(results)
    written: dict[str, str] = {}
    # The deleted-in-source finding is discovered by the runner into `context`; fold it into the
    # summary so every renderer below sees one object rather than needing both.
    summary = {**summary, "deleted_in_source": context.get("deleted_in_source", {})}

    # A live retry run tags its whole report set with a shared timestamp; dry + full-live keep their
    # existing names (dry only re-points the xlsx, see A1).
    _retry_active = (not summary.get("dry_run")
                     and safe_str(getattr(config.imports, "retry_mode", "off")) not in ("", "off"))
    _variant = f"retry_{datetime.now().strftime('%Y%m%d_%H%M%S')}" if _retry_active else ""

    # A retry report is SCOPED to the units the retry actually attempted: the units outside the
    # retry work list are flagged `retry_out_of_scope` and dropped here, so the retry file isn't the
    # whole inventory re-printed as "not outstanding" (the preserved full-run report already has
    # that). Totals/counts below then derive from the scoped rows, so the roll-up matches.
    if _retry_active:
        rows = [r for r in rows if not r.get("retry_out_of_scope")]
    totals = _rollup_totals(rows) if _retry_active else summary.get("totals", {})

    payload = {
        "run_id": summary.get("run_id"),
        "source_workspace_id": summary.get("source_workspace_id"),
        "connectivity_mode": summary.get("connectivity_mode"),
        "dry_run": summary.get("dry_run"),
        "run_status": summary.get("run_status"),
        "generated_utc": summary.get("generated_utc") or now_iso(),
        "elapsed_sec": summary.get("elapsed_sec"),
        "retry_mode": safe_str(getattr(config.imports, "retry_mode", "off")),
        "scoped_to_retry": _retry_active,
        "totals": totals,
        "per_phase": summary.get("per_phase", []),
        "counts_by_status": _counts_by_status(rows),
        "counts_by_asset_type": _counts_by_asset_type(rows),
        "deleted_in_source": context.get("deleted_in_source", {}),
        # The join surface Plan 4 reads: one row per unit, keyed (asset_type, natural_key).
        "units": rows,
    }
    # On a live retry the results json is tagged too, so the canonical (last full-run) join surface
    # is preserved; dry + full-live write the canonical name.
    json_rel = BP.with_variant(BP.IMPORT_RESULTS_JSON, _variant)
    try:
        aw.write_json(json_rel, payload)
        written["json"] = json_rel
    except Exception as exc:  # noqa: BLE001
        _LOG.warning("import results json not written", rel=json_rel, error=str(exc))

    # A1: a rehearsal writes a SEPARATE filename so a later live run never overwrites the dry-run's
    # report; a live retry writes `import_status_retry_<ts>.xlsx`. D-1: the ACL parity result is
    # folded in as a sheet, sourced from context.
    xlsx_rel = (BP.IMPORT_STATUS_DRYRUN_XLSX if summary.get("dry_run")
                else BP.with_variant(BP.IMPORT_STATUS_XLSX, _variant))
    parity = context.get("acl_parity") or {}
    # Bug 15: the ACL phase expands acls.json to one row per object×principal×permission; when
    # present, the workbook renders a dedicated "Object Permissions (ACLs)" sheet mirroring inventory
    # instead of collapsing each object to one `acl` row + a count.
    acl_grants = context.get("acl_grants")
    # PLAN 11 Finding-4: the cumulative outstanding items from the STATE TABLE (across all runs),
    # for the "Outstanding" sheet. Empty when state is disabled (a first-look dry run).
    outstanding = context.get("outstanding") or []
    try:
        path = aw.write_text_local_then_copy(
            xlsx_rel,
            lambda local: _render_xlsx(local, config, summary, rows, parity, acl_grants,
                                       outstanding))
        if path:
            written["xlsx"] = xlsx_rel
    except Exception as exc:  # noqa: BLE001 — Excel is a convenience; never fail the run
        _LOG.warning("import status xlsx not written", rel=xlsx_rel, error=str(exc))

    md_rel = BP.with_variant(BP.MANUAL_ACTIONS_IMPORT_MD, _variant)
    try:
        aw.write_bytes(md_rel, _render_manual_actions(summary, rows).encode("utf-8"))
        written["manual_actions"] = md_rel
    except Exception as exc:  # noqa: BLE001
        _LOG.warning("manual_actions_import.md not written", rel=md_rel, error=str(exc))

    _LOG.info("import reports written", **written)
    return written


def _counts_by_status(rows: list[dict]) -> dict:
    out: dict = {}
    for r in rows:
        s = safe_str(r.get("import_status"))
        out[s] = out.get(s, 0) + 1
    return out


# import_status → the ImportResult counter name, so a scoped retry report's `totals` matches the
# rows it actually shows (rather than the runner's whole-run counters).
_STATUS_TO_TOTAL = {
    "created": "created", "updated": "updated", "adopted": "adopted", "skipped": "skipped",
    "failed": "failed", "manual": "manual", "not_selected": "not_selected",
    "skipped_no_object": "skipped_no_object", "created_with_warning": "created_with_warning",
}


def _rollup_totals(rows: list[dict]) -> dict:
    """A totals dict (same shape as the runner's) computed from exactly the rows being reported —
    used for a retry report so its roll-up counts the scoped units, not the whole run."""
    out: dict = {"total": len(rows)}
    for r in rows:
        counter = _STATUS_TO_TOTAL.get(safe_str(r.get("import_status")))
        if counter:
            out[counter] = out.get(counter, 0) + 1
    return out


def _counts_by_asset_type(rows: list[dict]) -> dict:
    out: dict = {}
    for r in rows:
        at = safe_str(r.get("asset_type"))
        bucket = out.setdefault(at, {"total": 0})
        bucket["total"] += 1
        s = safe_str(r.get("import_status"))
        bucket[s] = bucket.get(s, 0) + 1
    return out



# ── Excel ───────────────────────────────────────────────────────────────────

def _render_xlsx(local_path: str, config, summary: dict, rows: list[dict],
                 parity: dict = None, acl_grants: list = None, outstanding: list = None) -> str:
    """Import Summary sheet + one sheet per asset type (with `deleted_in_source` shown INLINE,
    Finding-3) + a cumulative "Outstanding" sheet (Finding-4) + an "Object Permissions (ACLs)" sheet
    (per-grant, Bug 15) + an ACL Parity sheet (D-1). Rendered locally (openpyxl needs seeks)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Workspace Import — Status Summary"
    c.font = font(bold=True, color="FFFFFF", size=16)
    c.fill = fill("FF3621")
    c.alignment = centre
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = (f"Source workspace {summary.get('source_workspace_id')}   |   "
               f"Run {summary.get('run_id')}   |   Mode {summary.get('connectivity_mode')}   |   "
               f"{'DRY RUN — nothing written' if summary.get('dry_run') else 'LIVE'}   |   "
               f"Status {summary.get('run_status')}   |   "
               f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = font(italic=True, color="475569", size=9)
    c.alignment = centre

    # Roll-up by status
    r = 4
    ws.cell(row=r, column=1, value="Outcome roll-up").font = font(bold=True, size=12,
                                                                  color="1E3A5F")
    r += 1
    by_status = _counts_by_status(rows)
    for col, status in enumerate(_SUMMARY_ORDER, 1):
        label, colour = _STATUS_STYLE.get(status, (status, "FFFFFF"))
        h = ws.cell(row=r, column=col, value=label)
        h.font = font(bold=True, size=9)
        h.fill = fill(colour)
        h.border = box
        h.alignment = centre
        v = ws.cell(row=r + 1, column=col, value=by_status.get(status, 0))
        v.font = font(bold=True, size=12)
        v.border = box
        v.alignment = centre
    # Bug 4: a deletion is not an import_status on any PROCESSED unit, so it never appears in
    # `by_status` — give it its own roll-up cell so a deletion is visible at a glance rather than
    # only in the runbook.
    deleted = summary.get("deleted_in_source") or {}
    deleted_count = sum(len(v) for v in deleted.values())
    dcol = len(_SUMMARY_ORDER) + 1
    dh = ws.cell(row=r, column=dcol, value="Deleted in source")
    dh.font = font(bold=True, size=9)
    dh.fill = fill(_STATUS_STYLE["deleted_in_source"][1])
    dh.border = box
    dh.alignment = centre
    dv = ws.cell(row=r + 1, column=dcol, value=deleted_count)
    dv.font = font(bold=True, size=12)
    dv.border = box
    dv.alignment = centre
    r += 3

    # FAILURES FIRST — the whole point of the summary sheet is that nobody has to hunt for them.
    failures = [x for x in rows if safe_str(x.get("import_status")) == "failed"]
    ws.cell(row=r, column=1, value=f"Failures ({len(failures)})").font = font(
        bold=True, color="B91C1C", size=12)
    r += 1
    for col, h in enumerate(["Asset Type", "Natural Key", "Category",
                             "Actual server error + hint"], 1):
        cc = ws.cell(row=r, column=col, value=h)
        cc.font = font(bold=True, color="FFFFFF")
        cc.fill = fill("B91C1C")
        cc.border = box
    for x in failures:
        r += 1
        # Prefer the complete raw server error; fall back to the note (which already leads with the
        # server text) so the operator always sees what the target actually said — never a canned
        # message alone.
        reason = safe_str(x.get("error_raw")) or safe_str(x.get("note"))
        for col, v in enumerate([x.get("asset_type"), x.get("natural_key"),
                                 x.get("failure_category"), reason], 1):
            cc = ws.cell(row=r, column=col, value=safe_str(v))
            cc.fill = fill("FEE2E2")
            cc.border = box
            cc.font = font(size=9)
            cc.alignment = left_wrap
    r += 2

    # Manual actions table
    manual = [x for x in rows if safe_str(x.get("import_status")) == "manual"]
    ws.cell(row=r, column=1, value=f"Manual steps required ({len(manual)})").font = font(
        bold=True, color="92400E", size=12)
    r += 1
    for col, h in enumerate(["Asset Type", "Natural Key", "What a human must do"], 1):
        cc = ws.cell(row=r, column=col, value=h)
        cc.font = font(bold=True, color="FFFFFF")
        cc.fill = fill("92400E")
        cc.border = box
    for x in manual:
        r += 1
        for col, v in enumerate([x.get("asset_type"), x.get("natural_key"), x.get("note")], 1):
            cc = ws.cell(row=r, column=col, value=safe_str(v))
            cc.fill = fill("FEF3C7")
            cc.border = box
            cc.font = font(size=9)
            cc.alignment = left_wrap
    r += 2

    # Deleted-in-source table (Bug 4) — on target but no longer in the source bundle. NOT deleted;
    # deletion requires allow_deletes=true. Mirrors the runbook section so the xlsx is self-contained.
    if deleted:
        ws.cell(row=r, column=1,
                value=f"Deleted in source — review ({deleted_count})").font = font(
                    bold=True, color="9F1239", size=12)
        r += 1
        for col, h in enumerate(["Asset Type", "Natural Key", "Note"], 1):
            cc = ws.cell(row=r, column=col, value=h)
            cc.font = font(bold=True, color="FFFFFF")
            cc.fill = fill("9F1239")
            cc.border = box
        _del_note = ("on target but no longer in the source bundle — NOT deleted (set "
                     "allow_deletes=true to opt into deletion); confirm it was deliberately removed")
        for at in sorted(deleted):
            for key in sorted(deleted[at]):
                r += 1
                for col, v in enumerate([at, key, _del_note], 1):
                    cc = ws.cell(row=r, column=col, value=safe_str(v))
                    cc.fill = fill(_STATUS_STYLE["deleted_in_source"][1])
                    cc.border = box
                    cc.font = font(size=9)
                    cc.alignment = left_wrap
        r += 2

    # Per-asset-type counts
    ws.cell(row=r, column=1, value="Per asset type").font = font(bold=True, size=12,
                                                                 color="1E3A5F")
    r += 1
    headers = ["Asset Type", "Total"] + [_STATUS_STYLE[s][0] for s in _SUMMARY_ORDER]
    for col, h in enumerate(headers, 1):
        cc = ws.cell(row=r, column=col, value=h)
        cc.font = font(bold=True, color="FFFFFF", size=9)
        cc.fill = fill("1E3A5F")
        cc.border = box
        cc.alignment = centre
    by_type = _counts_by_asset_type(rows)
    for at in sorted(by_type):
        r += 1
        b = by_type[at]
        vals = [at, b["total"]] + [b.get(s, 0) for s in _SUMMARY_ORDER]
        for col, v in enumerate(vals, 1):
            cc = ws.cell(row=r, column=col, value=v)
            cc.border = box
            cc.font = font(size=9, bold=(col == 2))
            cc.alignment = centre if col > 1 else left_wrap

    # Bug 6: a purely account-level access-control change (group Manager, SP Can use / Can manage)
    # is invisible to this workspace-scoped tool, so such a row shows `Skipped (unchanged)` with no
    # targeted message possible. A one-time static disclaimer stops that being mistaken for a defect.
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    dc = ws.cell(row=r, column=1, value=_ACCOUNT_ACL_DISCLAIMER)
    dc.font = font(italic=True, color="475569", size=9)
    dc.alignment = left_wrap
    ws.row_dimensions[r].height = 56

    for i, w in enumerate([26, 52, 20, 70, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── one sheet PER ASSET TYPE, in inventory/export card order (IMP-1) ─────
    # Import keeps its own import-specific columns (status / action / ids / note), but the SHEET
    # LAYOUT — one tab per asset type, named + ordered like inventory.xlsx — now matches the other
    # two stages so an operator reads the same tabs across all three workbooks.
    from src.reports.inventory_view import _LABELS, _SUMMARY_CARD_KEYS

    cols = [("Asset Type", "asset_type", 24), ("Natural Key", "natural_key", 56),
            ("Import Status", "import_status", 22), ("Action Taken", "action_taken", 26),
            ("Target Id", "target_id", 24), ("Source Id", "source_id", 22),
            ("Failure Category", "failure_category", 20), ("Note / reason", "note", 70),
            ("Actual server error", "error_raw", 80)]

    by_card: dict[str, list] = {}
    for x in rows:
        by_card.setdefault(_card_for_asset_type(x.get("asset_type")), []).append(x)

    # PLAN 11 Finding-3: show `deleted_in_source` INLINE on each asset-type tab as a first-class
    # status (not only on the Summary sheet), so e.g. the Jobs tab reads created / updated / skipped
    # / deleted_in_source in one place. Synthetic rows are injected from the deleted map (sourced
    # from the state table's last_action=deleted_in_source), so a tab reflects them even on a run
    # that didn't otherwise touch that type. The Summary roll-up table keeps its own deleted section
    # (belt-and-suspenders); these injected rows are counted only in the per-type tab.
    _del_inline_note = ("on target but no longer in the source bundle — NOT deleted (set "
                        "allow_deletes=true to remove); a source rename shows as this old name "
                        "deleted + the new name created")
    for at, keys in (summary.get("deleted_in_source") or {}).items():
        for key in sorted(keys):
            by_card.setdefault(_card_for_asset_type(at), []).append({
                "asset_type": at, "natural_key": key, "import_status": "deleted_in_source",
                "action_taken": "Deleted in source", "note": _del_inline_note,
                "target_id": "", "source_id": "", "failure_category": "", "error_raw": ""})

    # Inventory card order first (so tabs line up with the other workbooks), then any leftover
    # cards not in the canonical list (defensive — a new importer type still gets a tab).
    ordered_cards = [k for k in _SUMMARY_CARD_KEYS if k in by_card]
    ordered_cards += [k for k in sorted(by_card) if k not in _SUMMARY_CARD_KEYS]
    # Bug 15: when per-grant ACL detail exists, the dedicated "Object Permissions (ACLs)" sheet
    # below replaces the collapsed one-row-per-object `object_permissions` tab — don't render both.
    if acl_grants is not None:
        ordered_cards = [k for k in ordered_cards if k != "object_permissions"]

    import re as _re
    for card in ordered_cards:
        label = _LABELS.get(card, card.replace("_", " ").title())
        sheet_name = _re.sub(r"[\\/?*\[\]:]", "-", label)[:31]
        sheet = wb.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        for col, (h, _k, w) in enumerate(cols, 1):
            cc = sheet.cell(row=1, column=col, value=h)
            cc.font = font(bold=True, color="FFFFFF", size=10)
            cc.fill = fill("1E3A5F")
            cc.border = box
            cc.alignment = centre
            sheet.column_dimensions[get_column_letter(col)].width = w
        for i, x in enumerate(by_card[card], start=2):
            status = safe_str(x.get("import_status"))
            _label, colour = _STATUS_STYLE.get(status, (status, "FFFFFF"))
            for col, (_h, key, _w) in enumerate(cols, 1):
                cc = sheet.cell(row=i, column=col, value=safe_str(x.get(key)))
                cc.border = box
                cc.font = font(size=9)
                cc.alignment = left_wrap
                if key == "import_status":
                    cc.fill = fill(colour)
                    cc.value = _label
        sheet.freeze_panes = "A2"

    # ── Outstanding — not yet successfully migrated (cumulative, from the STATE TABLE) ──────────
    # PLAN 11 Finding-4: driven from the state table for this pair (NOT this run's units), so it
    # always shows EVERY currently-unresolved item — old carry-overs AND new — whether or not this
    # run touched it. A persistent failure that stops being re-attempted (a stale row, exactly
    # BUG-1) would otherwise silently drop off future reports.
    if outstanding is not None:
        current_run = safe_str(summary.get("run_id"))
        osheet = wb.create_sheet("Outstanding")
        osheet.sheet_view.showGridLines = False
        # Scoped to genuine PROBLEMS only (customer 2026-09-04): failed + created_with_warning. The
        # state query (OUTSTANDING_ACTIONS) already excludes manual + skipped_no_object noise, but
        # count defensively here too so the banner never mislabels.
        _oa_counts = {"failed": 0, "created_with_warning": 0}
        for r in outstanding:
            a = safe_str(r.get("last_action"))
            if a in _oa_counts:
                _oa_counts[a] += 1
        total_out = sum(_oa_counts.values())
        osheet.merge_cells("A1:I1")
        c = osheet["A1"]
        c.value = (f"{total_out} outstanding: {_oa_counts['failed']} failed, "
                   f"{_oa_counts['created_with_warning']} created-with-warning")
        c.font = font(bold=True, color="FFFFFF", size=12)
        c.fill = fill("B91C1C" if _oa_counts["failed"] else "1E3A5F")
        c.alignment = centre
        osheet.row_dimensions[1].height = 26
        osheet.merge_cells("A2:I2")
        legend = osheet["A2"]
        legend.value = (
            "Cumulative PROBLEMS from the migration state table across ALL runs for this workspace "
            "pair — items that FAILED or were created-but-DEGRADED and still need a fix. failed = "
            "create/update errored; created_with_warning = created but a reference could not be "
            "fully resolved (fix the prerequisite + re-run with retry_mode=failed_only). Deliberately "
            "EXCLUDES routine by-design items so this stays scannable: manual steps (AKV scope, "
            "repos, secret values — see the Manual table + runbook), skipped_no_object ACLs (see the "
            "ACL sheet), up-to-date items (skipped/created/updated/adopted), deferred families "
            "(not_selected), and deletes (Summary sheet).")
        legend.font = font(italic=True, color="475569", size=9)
        legend.alignment = left_wrap
        osheet.row_dimensions[2].height = 66
        ocols = [("Asset Type", "asset_type", 24), ("Natural Key", "natural_key", 56),
                 ("Status", "last_action", 20), ("Origin", "_origin", 16),
                 ("Failure Category", "failure_category", 20), ("Last Error", "last_error", 70),
                 ("Last Run", "last_run_id", 20), ("First Seen", "first_seen", 26),
                 ("Last Seen", "last_seen", 26)]
        for col, (h, _k, w) in enumerate(ocols, 1):
            cc = osheet.cell(row=3, column=col, value=h)
            cc.font = font(bold=True, color="FFFFFF", size=10)
            cc.fill = fill("1E3A5F")
            cc.border = box
            cc.alignment = centre
            osheet.column_dimensions[get_column_letter(col)].width = w
        # Failures first, then by (asset_type, natural_key); newest-outstanding-first within is fine.
        ordered_out = sorted(
            outstanding,
            key=lambda r: (safe_str(r.get("last_action")) != "failed",
                           safe_str(r.get("asset_type")), safe_str(r.get("natural_key"))))
        for i, r in enumerate(ordered_out, start=4):
            status = safe_str(r.get("last_action"))
            _label, colour = _STATUS_STYLE.get(status, (status, "FFFFFF"))
            origin = ("new this run" if safe_str(r.get("last_run_id")) == current_run
                      else "carried over")
            values = {"asset_type": r.get("asset_type"), "natural_key": r.get("natural_key"),
                      "last_action": _label, "_origin": origin,
                      "failure_category": r.get("failure_category"), "last_error": r.get("last_error"),
                      "last_run_id": r.get("last_run_id"), "first_seen": r.get("first_seen"),
                      "last_seen": r.get("last_seen")}
            for col, (_h, key, _w) in enumerate(ocols, 1):
                cc = osheet.cell(row=i, column=col, value=safe_str(values.get(key)))
                cc.border = box
                cc.font = font(size=9)
                cc.alignment = left_wrap
                if key == "last_action":
                    cc.fill = fill(colour)
                elif key == "_origin":
                    cc.fill = fill("FEF3C7" if origin == "carried over" else "E0F2FE")
        osheet.freeze_panes = "A4"

    # ── Object Permissions (ACLs) — one row per object×principal×permission (Bug 15) ──
    # MIRRORS the inventory ACL sheet (Object Type · Object · Principal · Permission · Inherited) so
    # the three stages line up, and adds a per-grant Import Status + source/target id so an operator
    # can confirm a specific principal's grant on a specific object reached target WITHOUT hitting
    # the permissions API — the exact question the old collapsed "N grants applied" row could not
    # answer.
    if acl_grants is not None:
        aname = _re.sub(r"[\\/?*\[\]:]", "-",
                        _LABELS.get("object_permissions", "Object Permissions (ACLs)"))[:31]
        asheet = wb.create_sheet(aname)
        asheet.sheet_view.showGridLines = False
        acols = [("Object Type", "perm_object_type", 22), ("Object", "object", 52),
                 ("Principal", "principal", 40), ("Permission", "permission", 22),
                 ("Inherited", "inherited", 11), ("Import Status", "import_status", 30),
                 ("Source Id", "source_id", 22), ("Target Id", "target_id", 22)]
        for col, (h, _k, w) in enumerate(acols, 1):
            cc = asheet.cell(row=1, column=col, value=h)
            cc.font = font(bold=True, color="FFFFFF", size=10)
            cc.fill = fill("1E3A5F")
            cc.border = box
            cc.alignment = centre
            asheet.column_dimensions[get_column_letter(col)].width = w
        _grant_colour = {"applied": "D1FAE5", "failed": "FEE2E2",
                         "dropped — principal not on target": "FDE68A",
                         "skipped — no target object": "EDE9FE",
                         "skipped — inherited/built-in": "E5E7EB"}
        # Non-applied grants first (that is what needs an operator's eye), then by object + principal.
        ordered = sorted(acl_grants, key=lambda g: (
            safe_str(g.get("import_status")) == "applied",
            safe_str(g.get("perm_object_type")), safe_str(g.get("object")),
            safe_str(g.get("principal")), safe_str(g.get("permission"))))
        for i, g in enumerate(ordered, start=2):
            status = safe_str(g.get("import_status"))
            for col, (_h, key, _w) in enumerate(acols, 1):
                val = ("Yes" if g.get("inherited") else "No") if key == "inherited" else g.get(key)
                cc = asheet.cell(row=i, column=col, value=safe_str(val))
                cc.border = box
                cc.font = font(size=9)
                cc.alignment = left_wrap
                if key == "import_status":
                    cc.fill = fill(_grant_colour.get(status, "FFFFFF"))
        asheet.freeze_panes = "A2"

    # ── ACL Parity sheet (PLAN 7 §B2 / D-1) ──────────────────────────────────
    # Folded in from the ACL phase's post-apply diff — this is INDEPENDENT proof (every touched
    # target object was re-read and diffed against source), not the same thing as import status.
    # Only written when the acls family actually ran this session (parity is non-empty).
    if parity and parity.get("objects") is not None:
        psheet = wb.create_sheet("ACL Parity")
        psheet.sheet_view.showGridLines = False
        counts = parity.get("counts", {})
        psheet.merge_cells("A1:G1")
        c = psheet["A1"]
        c.value = (f"ACL parity — {parity.get('objects_checked', 0)} objects re-read and diffed "
                   f"against source   |   " + "   ".join(f"{k}={v}" for k, v in sorted(counts.items())))
        c.font = font(bold=True, color="FFFFFF", size=11)
        c.fill = fill("1E3A5F")
        c.alignment = left_wrap
        psheet.row_dimensions[1].height = 26
        # Bug 15: a `match` row now NAMES the principals verified present, not just diffs — so an
        # operator can confirm exactly which grants landed, on match rows too.
        pcols = [("Object Type", "perm_object_type", 22), ("Object", "object", 52),
                 ("Verdict", "verdict", 18), ("Verified present", "_present", 44),
                 ("Missing on target", "_missing", 40), ("Extra on target", "_extra", 40),
                 ("Detail", "detail", 46)]
        for col, (h, _k, w) in enumerate(pcols, 1):
            cc = psheet.cell(row=2, column=col, value=h)
            cc.font = font(bold=True, color="FFFFFF", size=10)
            cc.fill = fill("1E3A5F")
            cc.border = box
            cc.alignment = centre
            psheet.column_dimensions[get_column_letter(col)].width = w
        _verdict_colour = {"match": "D1FAE5", "missing_on_target": "FEE2E2",
                           "extra_on_target": "FDE68A", "both": "FECACA", "unverified": "E5E7EB"}
        objs = sorted(parity.get("objects", []),
                      key=lambda o: (safe_str(o.get("verdict")) == "match",
                                     safe_str(o.get("object"))))
        for i, o in enumerate(objs, start=3):
            verdict = safe_str(o.get("verdict"))
            missing = ", ".join(f"{m[0]}={m[1]}" for m in (o.get("missing_on_target") or []))
            extra = ", ".join(f"{e[0]}={e[1]}" for e in (o.get("extra_on_target") or []))
            present = ", ".join(f"{p[0]}={p[1]}" for p in (o.get("present") or []))
            values = {"perm_object_type": o.get("perm_object_type"), "object": o.get("object"),
                      "verdict": verdict, "_present": present, "_missing": missing, "_extra": extra,
                      "detail": o.get("detail", "")}
            for col, (_h, key, _w) in enumerate(pcols, 1):
                cc = psheet.cell(row=i, column=col, value=safe_str(values.get(key)))
                cc.border = box
                cc.font = font(size=9)
                cc.alignment = left_wrap
                if key == "verdict":
                    cc.fill = fill(_verdict_colour.get(verdict, "FFFFFF"))
        if parity.get("known_limitation"):
            note_row = len(objs) + 4
            psheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
            nc = psheet.cell(row=note_row, column=1, value=safe_str(parity.get("known_limitation")))
            nc.font = font(italic=True, color="475569", size=9)
            nc.alignment = left_wrap
        psheet.freeze_panes = "A3"

    wb.save(local_path)
    return local_path


# ── manual actions runbook ──────────────────────────────────────────────────

def _render_manual_actions(summary: dict, rows: list[dict]) -> str:
    """The import-side manual runbook.

    Written as its OWN file rather than overwriting export's `manual_actions.md`: the export-side
    rows (secret values to re-populate, DAB bundles to redeploy, oversize files to copy) still
    apply, and clobbering them would lose half the runbook.
    """
    manual = [r for r in rows if safe_str(r.get("import_status")) == "manual"]
    failed = [r for r in rows if safe_str(r.get("import_status")) == "failed"]
    warned = [r for r in rows if safe_str(r.get("import_status")) == "created_with_warning"]
    no_object = [r for r in rows if safe_str(r.get("import_status")) == "skipped_no_object"]

    out = [
        "# Manual actions after import",
        "",
        f"Run `{summary.get('run_id')}` — source workspace `{summary.get('source_workspace_id')}`"
        f"{' (DRY RUN — nothing was written)' if summary.get('dry_run') else ''}.",
        "",
        "This is the IMPORT-side runbook. The export-side `export/manual/manual_actions.md` still "
        "applies too (secret values, DAB redeploys, oversize files) — it is not superseded.",
        "",
    ]

    def section(title, items, guidance=""):
        if not items:
            return
        out.append(f"## {title} ({len(items)})")
        out.append("")
        if guidance:
            out.append(guidance)
            out.append("")
        by_type: dict[str, list] = {}
        for r in items:
            by_type.setdefault(safe_str(r.get("asset_type")), []).append(r)
        for at in sorted(by_type):
            out.append(f"### {at} ({len(by_type[at])})")
            for r in sorted(by_type[at], key=lambda x: safe_str(x.get("natural_key"))):
                note = safe_str(r.get("note"))
                out.append(f"- `{safe_str(r.get('natural_key'))}`" + (f" — {note}" if note else ""))
            out.append("")

    section("Manual recreate / re-populate", manual,
            "These have no REST create path in scope, so the tool never attempted them. Each line "
            "carries what it is and why.")
    section("Failures to fix, then retry", failed,
            "Fix the cause, then re-run `04_Import` with `retry_mode=failed_only` — that attempts "
            "ONLY these units and touches nothing else.")
    section("Created but degraded — verify before use", warned,
            "These EXIST on target but a reference could not be fully resolved (most often a "
            "notebook path inside a Git folder that has not been recreated). They will create "
            "fine and fail at first RUN, so they need a look. `retry_mode=failed_only` includes "
            "them once the prerequisite is in place.")
    section("Permissions not applied — target object absent", no_object,
            "The ACL could not be applied because its object does not exist on target yet "
            "(usually by design: DAB content, an out-of-scope repo, or a deferred family). Once "
            "the object exists, re-run with `import_assets=acls` and "
            "`retry_mode=skipped_only` to apply exactly these grants.")

    deleted = summary.get("deleted_in_source") or {}
    if deleted:
        out.append(f"## Deleted in source — review ({sum(len(v) for v in deleted.values())})")
        out.append("")
        out.append("These exist on TARGET (this tool created them on an earlier run) but are no "
                   "longer in the source bundle. They were **not** deleted — deletion requires "
                   "`allow_deletes=true`. Confirm each was deliberately removed on source.")
        out.append("")
        for at, keys in sorted(deleted.items()):
            out.append(f"### {at} ({len(keys)})")
            for k in sorted(keys):
                out.append(f"- `{k}`")
            out.append("")

    if len(out) <= 7:
        out.append("Nothing outstanding — no manual actions were recorded for this run.")

    # Bug 6: one-time static disclaimer so account-level access-control showing as `unchanged` is
    # not mistaken for a defect (a per-row note is impossible — the tool can't see the change).
    out += ["", "---", "## Note — account-level access-control (not tracked by this tool)", "",
            _ACCOUNT_ACL_DISCLAIMER]
    return "\n".join(out) + "\n"
