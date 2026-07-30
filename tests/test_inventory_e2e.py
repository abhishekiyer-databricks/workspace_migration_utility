"""End-to-end offline test: InventoryRunner → real artifacts in a temp dir.

Run: python3 -m tests.test_inventory_e2e
"""
from __future__ import annotations

import json
import os
import tempfile

from src.config.config_manager import Config
from src.collectors.inventory_runner import InventoryRunner
from src.exporters.artifact_writer import ArtifactWriter
from tests.fakes import FakeClient


def _client() -> FakeClient:
    def ws_list(params):
        if params.get("path") == "/":
            return {"objects": [{"path": "/Shared", "object_type": "DIRECTORY", "object_id": "1"}]}
        return {"objects": []}

    get_table = {
        "api/2.0/instance-pools/list": {"instance_pools": [{"instance_pool_id": "p1", "instance_pool_name": "pool"}]},
        "api/2.0/policies/clusters/list": {"policies": [{"policy_id": "pol", "name": "policy-1"}]},
        "api/2.0/clusters/list": {"clusters": [{"cluster_id": "c1", "cluster_name": "analytics", "cluster_source": "UI"}]},
        "api/2.0/secrets/scopes/list": {"scopes": [{"name": "kv", "backend_type": "AZURE_KEYVAULT",
                                                    "keyvault_metadata": {"dns_name": "u", "resource_id": "r"}}]},
        "api/2.0/secrets/acls/list": {"items": []},
        "api/2.0/secrets/list": {"secrets": [{"key": "k1"}]},
        "api/2.0/sql/warehouses": {"warehouses": [{"id": "w1", "name": "wh", "warehouse_type": "PRO"}]},
        "api/2.0/serving-endpoints": {"endpoints": []},
        "api/2.0/global-init-scripts": {"scripts": []},
        "api/2.0/libraries/all-cluster-statuses": {"statuses": []},
        "api/2.0/ip-access-lists": {"ip_access_lists": []},
        "api/2.0/workspace/list": ws_list,
        "api/2.0/workspace-conf": lambda p: {p["keys"]: "true"},
    }
    # ACL endpoints → empty ACL
    for t in ("instance-pools/p1", "cluster-policies/pol", "clusters/c1",
              "sql/warehouses/w1", "directories/1"):
        get_table[f"api/2.0/permissions/{t}"] = {"access_control_list": []}

    paginated = {
        "api/2.1/jobs/list": [{"job_id": "j1", "settings": {"name": "nightly", "format": "MULTI_TASK", "tasks": [{"task_key": "t"}]}}],
        "api/2.0/pipelines": [], "api/2.0/lakeview/dashboards": [], "api/2.0/genie/spaces": [],
        "api/2.0/repos": [], "api/2.0/sql/queries": [], "api/2.0/sql/alerts": [], "api/2.0/sql/dashboards": [],
    }
    scim = {
        "Users": [{"id": "u1", "userName": "alice@corp.com", "emails": [{"value": "alice@corp.com", "primary": True}],
                   "externalId": "ext", "entitlements": [{"value": "allow-cluster-create"}]}],
        "ServicePrincipals": [{"id": "s2", "applicationId": "app-dbx", "displayName": "dbx"}],  # DB-managed
        "Groups": [{"id": "g2", "displayName": "eng", "members": [{"value": "u1", "display": "alice@corp.com", "$ref": "Users/u1"}]}],
    }
    c = FakeClient(get_table=get_table, paginated_table=paginated, scim_table=scim)
    c.get_table["api/2.0/permissions/jobs/j1"] = {"access_control_list": [{"all_permissions": [{"permission_level": "IS_OWNER"}]}]}
    return c


def test_inventory_end_to_end():
    tmp = tempfile.mkdtemp()
    cfg = Config.from_dict({"role": "source", "source_workspace_id": "111", "run_id": "r1",
                            "source_staging_location": tmp})
    cfg.ctx.workspace_url = "https://adb-111.9.azuredatabricks.net"
    cfg.ctx.token = "SECRET-TOKEN"

    aw = ArtifactWriter(cfg)
    result = InventoryRunner(_client(), cfg, aw).run()
    root = result["output_path"]

    # JSON artifacts exist + parse
    inv = json.load(open(f"{root}/inventory.json"))
    assert inv["counts"]["identity"] == 3 and inv["counts"]["compute"] == 3
    idc = json.load(open(f"{root}/identity_classification.json"))
    assert idc["summary"].get("db_managed_sp") == 1 and idc["summary"].get("db_managed_group") == 1

    # config_resolved.json must NOT contain the token
    cr = json.load(open(f"{root}/config_resolved.json"))
    assert "SECRET-TOKEN" not in json.dumps(cr)

    # HTML exists and is the clickable reference-style app (sidebar + tabs + classification)
    html = open(f"{root}/inventory.html").read()
    assert "Workspace Inventory" in html and "Identity classification" in html and len(html) > 1000
    assert 'class="sidebar"' in html and "function showTab" in html  # reference app shell

    # Excel exists, non-zero, and opens (valid zip/xlsx) with the fine-grained sheets
    xlsx = f"{root}/inventory.xlsx"
    assert os.path.getsize(xlsx) > 0
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    assert {"Summary", "Migration Plan", "Users", "Groups", "Service Principals"} <= set(wb.sheetnames)

    # manifest verifies clean
    aw.write_manifest(result["counts"])
    assert aw.verify_manifest()["ok"]
    print("counts:", result["counts"])
    print("identity summary:", idc["summary"])


if __name__ == "__main__":
    import sys
    try:
        test_inventory_end_to_end()
        print("\nPASS  test_inventory_end_to_end")
    except Exception:
        import traceback
        traceback.print_exc()
        print("\nFAIL")
        sys.exit(1)
