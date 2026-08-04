"""End-to-end offline test: ExportRunner → a full bundle in a temp dir.

Runs inventory first (reusing the inventory e2e fake client), then Export, and asserts the whole
bundle: per-asset payload files, acls.json, content bytes, export_index.json reconciles 1:1
against inventory, export_status.xlsx opens, manifest verifies, and a resume re-run skips content.

Run: python3 -m tests.test_export_e2e
"""
from __future__ import annotations

import json
import os
import tempfile

from src.config.config_manager import Config
from src.collectors.inventory_runner import InventoryRunner
from src.exporters.artifact_writer import ArtifactWriter
from src.exporters.export_runner import ExportRunner
from tests.fakes import FakeClient


def _client() -> FakeClient:
    def ws_list(params):
        p = params.get("path")
        if p == "/":
            return {"objects": [{"path": "/Shared", "object_type": "DIRECTORY", "object_id": "1"},
                                {"path": "/Users/a@x.com", "object_type": "DIRECTORY", "object_id": "2"}]}
        if p == "/Shared":
            return {"objects": [{"path": "/Shared/nb", "object_type": "NOTEBOOK",
                                 "language": "PYTHON", "object_id": "10"}]}
        if p == "/Users/a@x.com":
            return {"objects": [{"path": "/Users/a@x.com/report", "object_type": "NOTEBOOK",
                                 "language": "SQL", "object_id": "11"},
                                {"path": "/Users/a@x.com/data.csv", "object_type": "FILE",
                                 "object_id": "12"}]}
        return {"objects": []}

    get_table = {
        "api/2.0/instance-pools/list": {"instance_pools": [{"instance_pool_id": "p1", "instance_pool_name": "pool"}]},
        "api/2.0/policies/clusters/list": {"policies": [{"policy_id": "pol", "name": "policy-1", "definition": "{}"}]},
        "api/2.0/clusters/list": {"clusters": [{"cluster_id": "c1", "cluster_name": "analytics",
                                                "cluster_source": "UI", "state": "RUNNING",
                                                "node_type_id": "n1", "spark_version": "13.x"}]},
        "api/2.0/secrets/scopes/list": {"scopes": [{"name": "kv", "backend_type": "AZURE_KEYVAULT",
                                                    "keyvault_metadata": {"dns_name": "u", "resource_id": "r"}}]},
        "api/2.0/secrets/acls/list": {"items": [{"principal": "eng", "permission": "MANAGE"}]},
        "api/2.0/secrets/list": {"secrets": [{"key": "k1"}, {"key": "k2"}]},
        "api/2.0/sql/warehouses": {"warehouses": [{"id": "w1", "name": "wh", "warehouse_type": "PRO",
                                                   "state": "RUNNING"}]},
        "api/2.0/serving-endpoints": {"endpoints": [
            {"name": "ext-ep", "id": "e1", "config": {"served_entities": [{"external_model": {"name": "gpt"}}]}}]},
        "api/2.0/global-init-scripts": {"scripts": [{"script_id": "g1", "name": "gis", "position": 0, "enabled": True}]},
        "api/2.0/global-init-scripts/g1": {"script": "ZWNobyBoaQ=="},
        "api/2.0/libraries/all-cluster-statuses": {"statuses": []},
        "api/2.0/workspace/list": ws_list,
        "api/2.0/workspace-conf": lambda p: {p["keys"]: "true"},
    }
    for t in ("instance-pools/p1", "cluster-policies/pol", "clusters/c1", "sql/warehouses/w1",
              "directories/1", "directories/2", "notebooks/10", "notebooks/11", "files/12",
              "serving-endpoints/e1"):
        get_table[f"api/2.0/permissions/{t}"] = {"access_control_list": []}
    # give the cluster a real ACL so acls.json is non-empty.
    get_table["api/2.0/permissions/clusters/c1"] = {"access_control_list": [
        {"user_name": "a@x.com", "all_permissions": [{"permission_level": "CAN_MANAGE"}]},
        {"group_name": "eng", "all_permissions": [{"permission_level": "CAN_ATTACH_TO"}]}]}

    paginated = {
        "api/2.1/jobs/list": [{"job_id": "j1", "settings": {"name": "nightly", "format": "MULTI_TASK",
                                                            "tasks": [{"task_key": "t"}]}}],
        "api/2.0/pipelines": [{"pipeline_id": "dp1", "name": "bronze"}],
        "api/2.0/lakeview/dashboards": [{"dashboard_id": "d1", "display_name": "KPIs"}],
        "api/2.0/genie/spaces": [{"space_id": "gs1", "title": "Genie", "warehouse_id": "w1"}],
        "api/2.0/repos": [], "api/2.0/sql/queries": [{"id": "q1", "name": "q"}],
        "api/2.0/sql/alerts": [], "api/2.0/sql/dashboards": [], "api/2.0/alerts": [],
        "api/2.0/apps": [], "api/2.0/postgres/projects": [],
    }
    get_table["api/2.0/pipelines/dp1"] = {"spec": {"name": "bronze", "target": "db"}}
    get_table["api/2.0/lakeview/dashboards/d1"] = {"display_name": "KPIs", "warehouse_id": "w1",
                                                   "serialized_dashboard": "{}"}
    # Genie space detail with serialized_space (verified live endpoint shape) → auto-migratable.
    get_table["api/2.0/genie/spaces/gs1"] = {
        "space_id": "gs1", "title": "Genie", "description": "d", "warehouse_id": "w1",
        "serialized_space": '{"version":2,"data_sources":{"tables":[{"identifier":"c.s.t"}]}}'}
    get_table["api/2.0/permissions/pipelines/dp1"] = {"access_control_list": []}
    get_table["api/2.0/permissions/jobs/j1"] = {"access_control_list": [
        {"all_permissions": [{"permission_level": "IS_OWNER"}]}]}
    get_table["api/2.0/permissions/queries/q1"] = {"access_control_list": []}
    get_table["api/2.0/permissions/genie/gs1"] = {"access_control_list": []}

    scim = {
        "Users": [{"id": "u1", "userName": "a@x.com", "emails": [{"value": "a@x.com", "primary": True}],
                   "externalId": "ext", "entitlements": [{"value": "allow-cluster-create"}]}],
        "ServicePrincipals": [{"id": "s2", "applicationId": "app-dbx", "displayName": "dbx"}],
        "Groups": [{"id": "g2", "displayName": "eng",
                    "members": [{"value": "u1", "display": "a@x.com", "$ref": "Users/u1"}]}],
    }
    download = {
        "api/2.0/workspace/export": lambda p: (b"-- Databricks notebook source\nSELECT 1\n"
                                               if p.get("path", "").endswith("report")
                                               else b"# Databricks notebook source\nprint('hi')\n"
                                               if "nb" in p.get("path", "")
                                               else b"col1,col2\n1,2\n"),
    }
    return FakeClient(get_table=get_table, paginated_table=paginated, scim_table=scim,
                      download_table=download)


def test_export_end_to_end():
    tmp = tempfile.mkdtemp()
    cfg = Config.from_dict({"role": "source", "source_workspace_id": "111", "run_id": "r1",
                            "source_staging_location": tmp})
    cfg.ctx.workspace_url = "https://adb-111.azuredatabricks.net"
    cfg.ctx.token = "SECRET"
    client = _client()

    aw = ArtifactWriter(cfg)
    InventoryRunner(client, cfg, aw).run()
    inv = aw.read_json("inventory.json")

    result = ExportRunner(client, cfg, aw, content_fetch_workers=4).run()
    root = result["output_path"]

    # ── export_index.json exists + reconciles 1:1 with inventory unit counts ──
    index = json.load(open(f"{root}/export_index.json"))
    units = index["units"]
    # every unit has a status + fingerprint + natural_key.
    assert all(u["export_status"] and u["fingerprint"].startswith("sha256:") and u["natural_key"]
               for u in units)
    by_type = {}
    for u in units:
        by_type.setdefault(u["asset_type"], []).append(u)
    # identity: 1 user + 1 sp + 1 group = 3.
    assert len(by_type["user"]) == 1 and len(by_type["service_principal"]) == 1
    assert len(by_type["group"]) == 1
    # workspace: 2 dirs + 2 notebooks + 1 file.
    assert len(by_type["notebook"]) == 2 and len(by_type["workspace_file"]) == 1
    assert len(by_type["directory"]) == 2
    # secret_scope 1 + 2 secret_value keys.
    assert len(by_type["secret_scope"]) == 1 and len(by_type["secret_value"]) == 2
    # Genie is now AUTO-migratable (serialized_space captured); payload carries it.
    assert by_type["genie_space"][0]["export_status"] == "success"
    genie_payload = json.load(open(f"{root}/export/genie/spaces.json"))["units"][0]["payload"]
    assert genie_payload["serialized_space"] and genie_payload["warehouse_id"] == "w1"

    # ── per-asset payload files exist + are create-ready ──
    users = json.load(open(f"{root}/export/identity/users.json"))
    assert users["units"][0]["payload"]["userName"] == "a@x.com"
    # The importer reads these per-asset files, so an identity unit must carry BOTH its
    # classification and the derived import_action here — not only in export_index.json.
    # (Regression: import_action was stamped on the index but dropped from the payload file,
    # so an account-managed SP would have arrived at the importer with no create-vs-assign
    # instruction; creating one instead of assigning mints a new appId and orphans its ACLs.)
    for u in users["units"]:
        assert u.get("classification"), f"classification missing on {u['natural_key']}"
        assert u.get("import_action"), f"import_action missing on {u['natural_key']}"
    clusters = json.load(open(f"{root}/export/compute/clusters.json"))
    assert "state" not in clusters["units"][0]["payload"]   # runtime stripped
    assert clusters["units"][0]["payload"]["node_type_id"] == "n1"
    jobs = json.load(open(f"{root}/export/jobs.json"))
    assert jobs["units"][0]["payload"]["tasks"]

    # ── acls.json separate + cluster grants captured, principals verbatim ──
    acls = json.load(open(f"{root}/export/acls.json"))
    cl_acl = next(e for e in acls if e["asset_type"] == "cluster")
    assert cl_acl["perm_object_type"] == "clusters" and len(cl_acl["grants"]) == 2
    # the cluster unit's acl_grants count was stamped.
    assert by_type["cluster"][0]["acl_grants"] == 2

    # ── content bytes written + content_ref set + correct extension ──
    content_units = by_type["notebook"] + by_type["workspace_file"]
    for u in content_units:
        assert u["export_status"] == "success" and u["content_ref"]
        full = os.path.join(root, u["content_ref"])
        assert os.path.isfile(full) and os.path.getsize(full) > 0
    exts = {os.path.splitext(u["content_ref"])[1] for u in by_type["notebook"]}
    assert exts == {".py", ".sql"}
    assert by_type["workspace_file"][0]["content_ref"].endswith(".bin")

    # ── manual_actions.md + excel + manifest ──
    assert os.path.isfile(f"{root}/export/manual/manual_actions.md")
    md = open(f"{root}/export/manual/manual_actions.md").read()
    assert "secret_value" in md   # secret values are always manual
    xlsx = f"{root}/export_status.xlsx"
    assert os.path.getsize(xlsx) > 0
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    assert "Export Summary" in wb.sheetnames
    # a per-asset sheet carries the Export Status column.
    users_sheet = wb["Users"]
    header = [users_sheet.cell(row=3, column=col).value for col in range(1, users_sheet.max_column + 1)]
    assert "Export Status" in header

    # config_resolved.json got export_options AND still no token.
    cr = json.load(open(f"{root}/config_resolved.json"))
    assert cr["export_options"]["content_fetch_workers"] == 4
    assert "SECRET" not in json.dumps(cr)

    # manifest verifies clean (written last → bundle complete).
    assert aw.verify_manifest()["ok"]

    # ── resume: a second run skips already-fetched content (checkpoint) ──
    # Break the download so a re-fetch WOULD fail; resume must not re-fetch.
    client.download_table = {}
    aw2 = ArtifactWriter(cfg)
    result2 = ExportRunner(client, cfg, aw2, content_fetch_workers=4).run()
    index2 = json.load(open(f"{root}/export_index.json"))
    nb2 = [u for u in index2["units"] if u["asset_type"] == "notebook"]
    assert all(u["export_status"] == "success" for u in nb2), "resume must reuse prior content"
    assert result2["failure"] == 0

    print("counts:", {k: len(v) for k, v in by_type.items()})
    print("summary:", {k: result[k] for k in ("total", "success", "manual", "dab", "skip")})


if __name__ == "__main__":
    import sys
    try:
        test_export_end_to_end()
        print("\nPASS  test_export_end_to_end")
    except Exception:
        import traceback
        traceback.print_exc()
        print("\nFAIL")
        sys.exit(1)
