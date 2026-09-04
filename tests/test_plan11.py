"""PLAN 11 regression tests — one (or more) per finding, so each fix is locked against re-drift.

Findings covered here (others are locked in the phase tests they naturally belong to):
  BUG-1     — the create-race adopt path HEALS a stale object instead of stranding it
  Finding-9 — full-path natural keys (no same-name collapse) + legacy_query UPDATE via PATCH
  Finding-10— exact-or-fail-loud reference remap: jobs sql_task/pipeline_task/run_job_task; run_as
  Finding-8 — orphaned-owner divert to the backup root for folder-placed assets
  Finding-7 — SP OAuth-secret manual unit is KIND-scoped (locked in test_identity_importer.py too)
  Finding-2 — account/Entra group membership note names the delta + the account-managed clause
  Finding-4 — cumulative Outstanding sheet driven from the state table
  Finding-3 — deleted_in_source shown inline on each asset-type tab
  Finding-12— configurable DAB bundle roots (glob OR directory prefix), default byte-identical
"""
from __future__ import annotations

import json
import os
import tempfile

from src.config.config_manager import Config
from src.exporters import bundle_paths as BP
from src.exporters.artifact_writer import ArtifactWriter
from src.importers.base_importer import BaseImporter
from src.importers.identity_importer import IdentityImporter
from src.importers.jobs_importer import JobsImporter
from src.importers.sql_importer import SqlImporter
from src.state.state_store import StateStore
from src.utils.helpers import dab_path_info, folder_natural_key, is_bundle_root_path
from tests.test_importers_phase2_5 import RecordingClient, _make, _unit
from tests.test_state_store import FakeBackend


# ═══════════════════════════ Finding-9 — no same-name collapse ══════════════════════════════

def test_two_same_named_queries_in_different_folders_both_create():
    """Finding-9 Bug B: two DISTINCT queries both named "New query" in DIFFERENT folders must
    migrate as TWO target objects — never collapse onto one (N-1 silently lost)."""
    client = RecordingClient()
    imp, _st = _make(SqlImporter, [
        _unit("legacy_query", "/Users/a@x.com/New query",
              {"display_name": "New query", "query_text": "select 1"}),
        _unit("legacy_query", "/Users/b@x.com/New query",
              {"display_name": "New query", "query_text": "select 2"}),
    ], client)
    res = imp.run()
    creates = client.bodies_to("sql/queries")
    assert res.created == 2, "both same-named queries must create as distinct objects"
    assert len(creates) == 2
    # distinct state keys → distinct target ids
    tids = set((imp.context.get("legacy_query_target_ids") or {}).values())
    assert len(tids) == 2


def test_a_changed_query_updates_via_PATCH_on_the_right_target_id():
    """Finding-9 Bug A: legacy_query UPDATE is a PATCH on the modern Queries API (POST-with-id 404s
    ENDPOINT_NOT_FOUND), hitting the target id resolved from STATE (id-anchor), not a name map."""
    client = RecordingClient(paginated={
        "api/2.0/sql/queries": [{"display_name": "New query", "id": "Q1"}],
        "api/2.0/alerts": []})
    key = "/Users/a@x.com/New query"
    imp, st = _make(SqlImporter, [
        _unit("legacy_query", key, {"display_name": "New query", "query_text": "select 99"},
              fingerprint="sha256:new")], client)
    st.record("legacy_query", key, action="created", fingerprint="sha256:old",
              target_object_id="Q1")
    res = imp.run()
    patches = [c for c in client.calls if c[0] == "PATCH" and "sql/queries/Q1" in c[1]]
    assert patches, "a changed query must PATCH /api/2.0/sql/queries/{id}"
    assert "update_mask" in patches[0][2], "the modern Queries PATCH needs an update_mask"
    assert client.posts_to("sql/queries") == [], "an UPDATE must not POST a duplicate"
    assert res.updated == 1


def test_folder_natural_key_helper():
    assert folder_natural_key("/Workspace/Users/a", "New query") == "/Users/a/New query"
    assert folder_natural_key("", "solo") == "solo"
    assert folder_natural_key(None, "solo") == "solo"


# ═══════════════════════════ BUG-1 — create-race adopt heals ════════════════════════════════

class _RacyImporter(BaseImporter):
    """Toy importer: existence map MISSES the object, create raises RESOURCE_ALREADY_EXISTS."""
    component = "sql"
    asset_types = ("alert_v2",)

    def __init__(self, *a, **kw):
        super().__init__(*a, **kw)
        self.updated_with = []

    def load(self):
        return self.units_for("alert_v2")

    def existing_keys(self):
        return {}                      # the map MISSES it — exactly the BUG-1 condition

    def create_one(self, unit):
        raise RuntimeError("RESOURCE_ALREADY_EXISTS: an alert with that name exists")

    def update_one(self, unit, target_id):
        self.updated_with.append((self.natural_key(unit), target_id))
        return {"target_id": target_id, "note": "updated in place"}


def test_create_race_adopt_heals_a_stale_object_instead_of_stranding_it():
    """BUG-1: the decision resolved to CREATE (existence map missed), the create raced
    RESOURCE_ALREADY_EXISTS, AND the stored fingerprint moved → the object must be UPDATED (healed
    against the state-stored target id), not recorded `adopted` with the edit silently dropped."""
    client = RecordingClient()
    key = "/Users/a@x.com/wsmig_test_alert_v2"
    imp, st = _make(_RacyImporter, [
        _unit("alert_v2", key, {"display_name": "wsmig_test_alert_v2"},
              fingerprint="sha256:NEW")], client)
    st.record("alert_v2", key, action="adopted", fingerprint="sha256:OLD",
              target_object_id="4308521176234470")
    res = imp.run()
    assert imp.updated_with == [(key, "4308521176234470")], \
        "the create-race adopt path must resolve the real target id and UPDATE it"
    assert res.updated == 1 and res.adopted == 0
    assert st.row("alert_v2", key)["last_action"] == "updated"
    assert st.row("alert_v2", key)["last_source_fingerprint"] == "sha256:NEW", \
        "the fingerprint may advance only AFTER the update lands"


def test_create_race_adopt_stays_adopt_when_fingerprint_unchanged():
    """The heal only fires on a MOVED fingerprint — an unchanged one is a plain adopt (no wasted
    update call)."""
    client = RecordingClient()
    key = "/Users/a@x.com/alert"
    imp, st = _make(_RacyImporter, [
        _unit("alert_v2", key, {"display_name": "alert"}, fingerprint="sha256:SAME")], client)
    st.record("alert_v2", key, action="created", fingerprint="sha256:SAME",
              target_object_id="T1")
    res = imp.run()
    assert imp.updated_with == [], "unchanged fingerprint → no update"
    assert res.adopted == 1


# ═══════════════════════════ Finding-10 — jobs task references ══════════════════════════════

def _job(tasks):
    return _unit("job", "etl-job", {"name": "etl-job", "tasks": tasks})


def test_job_sql_task_warehouse_is_remapped_to_the_target():
    client = RecordingClient()
    imp, _st = _make(JobsImporter, [_job([
        {"task_key": "q", "sql_task": {"warehouse_id": "SRC-WH", "query": {"query_id": "x"}}}])],
        client, context={"sql_warehouse_target_ids": {"wh": "TGT-WH"}})
    imp.units_by_type["sql_warehouse"] = [_unit("sql_warehouse", "wh", {"name": "wh"},
                                                source_id="SRC-WH")]
    imp.run()
    body = client.bodies_to("2.1/jobs/create") or client.bodies_to("jobs/create")
    task = body[0]["tasks"][0]
    assert task["sql_task"]["warehouse_id"] == "TGT-WH", "sql_task.warehouse_id must be remapped"


def test_job_pipeline_and_run_job_tasks_are_remapped():
    client = RecordingClient()
    imp, _st = _make(JobsImporter, [_job([
        {"task_key": "p", "pipeline_task": {"pipeline_id": "SRC-PIPE"}},
        {"task_key": "j", "run_job_task": {"job_id": "SRC-JOB"}}])], client,
        context={"dlt_pipeline_target_ids": {"pipe": "TGT-PIPE"},
                 "job_target_ids": {"other": "TGT-JOB"}})
    imp.units_by_type["dlt_pipeline"] = [_unit("dlt_pipeline", "pipe", {"name": "pipe"},
                                               source_id="SRC-PIPE")]
    imp.units_by_type["job"] = list(imp.units_by_type.get("job", [])) + [
        _unit("job", "other", {"name": "other"}, source_id="SRC-JOB")]
    imp.run()
    body = client.bodies_to("2.1/jobs/create") or client.bodies_to("jobs/create")
    tasks = {t["task_key"]: t for t in body[0]["tasks"]}
    assert tasks["p"]["pipeline_task"]["pipeline_id"] == "TGT-PIPE"
    assert tasks["j"]["run_job_task"]["job_id"] == "TGT-JOB"


def test_job_task_reference_not_in_bundle_fails_loud():
    """Finding-10: a task warehouse that is NOT in the bundle is a HARD failure, never left as the
    source id."""
    client = RecordingClient()
    imp, st = _make(JobsImporter, [_job([
        {"task_key": "q", "sql_task": {"warehouse_id": "GONE"}}])], client)
    res = imp.run()
    assert res.created == 0 and res.failed == 1
    row = st.row("job", "etl-job")
    assert row["failure_category"] == "dependency_unresolved"
    assert "not available on source" in row["last_error"]


def test_run_as_account_sp_not_in_map_is_left_as_is_not_failed():
    """Finding-10: run_as is the ONE correct exception — an account SP's appId is stable, so an
    unmapped run_as is LEFT AS-IS (a warning), never a hard fail."""
    client = RecordingClient()
    imp, _st = _make(JobsImporter, [
        _unit("job", "j", {"name": "j", "tasks": [],
                           "run_as": {"service_principal_name": "acct-app-id"}})], client)
    res = imp.run()
    body = client.bodies_to("2.1/jobs/create") or client.bodies_to("jobs/create")
    assert body[0]["run_as"]["service_principal_name"] == "acct-app-id", "run_as left as-is"
    # created (possibly created_with_warning for the unmapped-run_as note) but NEVER a hard failure
    assert res.failed == 0 and (res.created + res.warned) == 1


# ═══════════════════════════ Finding-8 — orphan-owner divert ════════════════════════════════

def _roster_file(usernames):
    return {BP.IDENTITY_CLASSIFICATION_JSON: json.dumps(
        {"identities": [{"identity_type": "user", "userName": u} for u in usernames]}
    ).encode("utf-8")}


def test_orphaned_owner_query_is_diverted_to_backup_as_created_with_warning():
    """Finding-8: a query owned by a user ABSENT from the roster (deleted in source) is PRESERVED
    under the backup root as created_with_warning — parity with notebooks, not a hard failure."""
    client = RecordingClient()   # get-status 404s → the home is absent on target
    imp, _st = _make(SqlImporter, [
        _unit("legacy_query", "/Users/gone@x.com/Drafts/New query",
              {"display_name": "New query", "parent_path": "/Users/gone@x.com/Drafts"})],
        client, staging_files=_roster_file(["present@x.com"]))
    res = imp.run()
    assert res.warned == 1 and res.failed == 0
    body = client.bodies_to("sql/queries")[0]["query"]
    assert body["parent_path"] == "/Users_Backup/gone@x.com/Drafts", \
        "the query must land under the backup root"
    row = next(r for r in res.units if r["asset_type"] == "legacy_query")
    assert "deleted in source" in row["note"] and "preserved" in row["note"]
    # the backup folder is provisioned (create APIs do not auto-mkdir a parent)
    assert any(c[0] == "POST" and "workspace/mkdirs" in c[1]
               and c[2].get("path") == "/Users_Backup/gone@x.com/Drafts" for c in client.calls)


def test_in_roster_owner_is_not_diverted_to_backup():
    """Finding-8: an owner still IN the roster (home just not present yet) must NOT be silently
    diverted — the source path is kept so it recovers into the real home on retry."""
    client = RecordingClient()
    imp, _st = _make(SqlImporter, [
        _unit("legacy_query", "/Users/present@x.com/Drafts/q",
              {"display_name": "q", "parent_path": "/Users/present@x.com/Drafts"})],
        client, staging_files=_roster_file(["present@x.com"]))
    imp.run()
    body = client.bodies_to("sql/queries")[0]["query"]
    assert body["parent_path"] == "/Users/present@x.com/Drafts", \
        "an in-roster owner must not be diverted to /Users_Backup"


# ═══════════════════════════ Finding-2 — account-group membership note ══════════════════════

def test_account_group_membership_change_names_the_delta_and_account_managed_clause():
    """Finding-2: a membership-only change on an ACCOUNT group is `updated` (fingerprint moved) and
    the note NAMES the member delta + the account-managed clause — never the false 'no source-side
    change detected'."""
    client = RecordingClient()
    imp, st = _make(IdentityImporter, [], client)
    st.record("group", "acct-grp", action="adopted", fingerprint="old", target_object_id="G1",
              source_detail=json.dumps({"entitlements": [], "roles": [],
                                        "members": ["Alice"]}, sort_keys=True))
    st.flush()
    unit = {"asset_type": "group", "natural_key": "acct-grp", "kind": "account",
            "payload": {"displayName": "acct-grp",
                        "members": [{"display": "Alice"}, {"display": "Bob"}]}}
    out = imp.update_one(unit, "G1")
    assert "membership changed in source" in out["note"]
    assert "Bob" in out["note"] and "account-managed" in out["note"]
    assert "no source-side change detected" not in out["note"]


# ═══════════════════════════ Finding-4 / Finding-3 — report sheets ══════════════════════════

def _render(rows, outstanding=None, deleted=None, run_id="RUN-2"):
    from src.reports.import_report import _render_xlsx
    import openpyxl
    cfg = Config.from_dict({"role": "target", "source_workspace_id": "111", "run_id": run_id,
                            "target_staging_location": "/tmp"})
    summary = {"run_id": run_id, "source_workspace_id": "111", "connectivity_mode": "direct",
               "dry_run": False, "run_status": "completed",
               "deleted_in_source": deleted or {}}
    path = os.path.join(tempfile.mkdtemp(), "s.xlsx")
    _render_xlsx(path, cfg, summary, rows, None, None, outstanding)
    return openpyxl.load_workbook(path)


def test_outstanding_sheet_is_driven_from_state_with_origin_column():
    """Finding-4: a cumulative Outstanding sheet, sourced from the state table (not this run's
    units), with an Origin column (new this run vs carried over) and the totals banner."""
    outstanding = [
        {"asset_type": "job", "natural_key": "j1", "last_action": "failed",
         "failure_category": "api_error", "last_error": "boom", "last_run_id": "RUN-2",
         "first_seen_utc": "t0", "last_seen_utc": "t2"},
        {"asset_type": "cluster_library", "natural_key": "c:lib", "last_action": "manual",
         "failure_category": "", "last_error": "hand-install", "last_run_id": "RUN-1",
         "first_seen_utc": "t0", "last_seen_utc": "t1"},
    ]
    wb = _render([], outstanding=outstanding)
    assert "Outstanding" in wb.sheetnames
    ws = wb["Outstanding"]
    text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "2 outstanding" in text and "1 failed" in text and "1 manual" in text
    assert "new this run" in text and "carried over" in text
    assert "j1" in text and "c:lib" in text


def test_deleted_in_source_shows_inline_on_the_asset_type_tab():
    """Finding-3: a `deleted_in_source` item appears on its asset-type tab (the Jobs tab), not only
    on the Summary sheet."""
    rows = [{"asset_type": "job", "natural_key": "live-job", "import_status": "created",
             "action_taken": "Created on target", "target_id": "T1", "source_id": "S1",
             "note": "", "failure_category": "", "error_raw": ""}]
    wb = _render(rows, deleted={"job": ["old-job"]})
    # the Jobs tab (card label "Jobs")
    jobs_sheet = next(wb[n] for n in wb.sheetnames if n.lower().startswith("job"))
    text = "\n".join(str(c.value) for row in jobs_sheet.iter_rows() for c in row if c.value)
    assert "old-job" in text, "the deleted-in-source job must appear on the Jobs tab"
    assert "Deleted in source" in text


# ═══════════════════════════ Finding-12 — configurable DAB roots ════════════════════════════

def test_dab_path_info_default_is_unchanged():
    assert dab_path_info("/Shared/.bundle/b/x") == {
        "deployed_by_dab": True, "dab_scope": "shared", "bundle_root": "/Shared/.bundle"}
    assert dab_path_info("/Users/u@x/.bundle/b/x")["dab_scope"] == "user"
    assert dab_path_info("/Workspace/Shared/.bundle/b/x")["dab_scope"] == "shared"
    assert dab_path_info("/Shared/regular/nb")["deployed_by_dab"] is False
    # the .bundle folder ITSELF (last segment) is not "content under a root"
    assert dab_path_info("/Shared/.bundle")["deployed_by_dab"] is False


def test_dab_path_info_directory_prefix_root():
    roots = ["/Users/dab@corp.com/prod"]
    assert dab_path_info("/Users/dab@corp.com/prod/dash", roots)["deployed_by_dab"] is True
    assert dab_path_info("/Users/dab@corp.com/prod/dash", roots)["dab_scope"] == "user"
    assert dab_path_info("/Users/dab@corp.com/other/dash", roots)["deployed_by_dab"] is False
    # with ONLY a directory root, .bundle no longer matches
    assert dab_path_info("/Shared/.bundle/b/x", roots)["deployed_by_dab"] is False


def test_dab_path_info_glob_root_and_mixed():
    assert dab_path_info("/Shared/myteam.bundle/b/x", ["*.bundle"])["deployed_by_dab"] is True
    # a mixed workspace hosting both conventions
    both = [".bundle", "/Users/dab@corp.com"]
    assert dab_path_info("/Shared/.bundle/b/x", both)["deployed_by_dab"] is True
    assert dab_path_info("/Users/dab@corp.com/anything", both)["deployed_by_dab"] is True


def test_is_bundle_root_path_for_state_file_discovery():
    roots = ["/Users/dab@corp.com/prod"]
    assert is_bundle_root_path("/Users/dab@corp.com/prod/state/resources.json", roots) is True
    assert is_bundle_root_path("/Shared/.bundle/b/state/resources.json") is True
