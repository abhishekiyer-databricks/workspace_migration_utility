"""Offline tests for the preflight gate and the import report set (Plan 3 §1a, §9, D8, D14, D16).

Preflight's value is entirely in its GRADING: a check that blocks when it should only warn stops a
migration for no reason, and one that warns when it should block lets a wrong target through. So the
grade of each finding is what's asserted, not just pass/fail.

The report tests assert the two things the customer will actually rely on: that `import_results.json`
is joinable on `(asset_type, natural_key)` (which is all Plan 4 needs from import), and that failures
are impossible to miss.
"""
from __future__ import annotations

import json
import os
import tempfile

from src.config.config_manager import Config
from src.exporters import bundle_paths as BP
from src.exporters.artifact_writer import ArtifactWriter
from src.importers.base_importer import ImportResult
from src.importers.preflight import BLOCKING, DEGRADING, GO, GO_WITH_WARNINGS, NO_GO, Preflight
from src.reports.import_report import write_import_reports
from src.state.state_store import StateStore
from tests.test_importers_phase2_5 import RecordingClient
from tests.test_state_store import FakeBackend


def _bundle(*, dry_run=True, state=True, files=None, **cfg_over):
    """A valid bundle + config, plus a working state store when `state` (the normal case)."""
    d = tempfile.mkdtemp()
    conf = {"role": "target", "source_workspace_id": "111", "run_id": "r1",
            "target_staging_location": d, "dry_run": dry_run,
            "imports": ({"state_catalog": "c", "state_schema": "s"} if state else {})}
    conf.update(cfg_over)
    cfg = Config.from_dict(conf)
    aw = ArtifactWriter(cfg)
    aw.ensure_output_path()
    for rel, data in (files or {}).items():
        aw.write_json(rel, data)
    # A valid manifest for everything written so far, so the bundle check passes by default.
    aw.write_manifest({})
    return cfg, aw


def _store(cfg):
    """A reachable state store — preflight treats an ABSENT one as BLOCKING, which is correct but
    would mask the finding under test in most of these cases."""
    if not cfg.state_enabled:
        return None
    st = StateStore(FakeBackend(), cfg)
    return st


def _scim_client(users=(), sps=(), groups=(), warehouses=(("wh-1", "target-wh"),),
                 account_groups=()):
    """`groups` are workspace-local on target; `account_groups` are assigned account groups.

    The distinction matters: a workspace-local group holding an ACCOUNT group's name is the shadow
    that permanently blocks assignment (Plan 6 F6), so preflight must tell them apart via
    `meta.resourceType` exactly as the live API does.
    """
    client = RecordingClient(get_table={
        "api/2.0/sql/warehouses": {"warehouses": [{"id": i, "name": n} for i, n in warehouses]},
        "scim:Users": [{"userName": u} for u in users],
        "scim:ServicePrincipals": [{"applicationId": s} for s in sps],
        "scim:Groups": ([{"displayName": g, "meta": {"resourceType": "WorkspaceGroup"}}
                         for g in groups]
                        + [{"displayName": g, "meta": {"resourceType": "Group"}}
                           for g in account_groups]),
    })
    return client


def _find(report, check_substring):
    return next(f for f in report["findings"] if check_substring in f["check"])


# ── verdicts + grading ─────────────────────────────────────────────────────

def test_a_clean_bundle_on_a_healthy_target_is_GO():
    cfg, aw = _bundle(files={BP.IDENTITY_CLASSIFICATION_JSON: {"identities": []},
                             BP.EXPORT_INDEX_JSON: {"units": []}})
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    assert report["verdict"] == GO, f"unexpected blockers/warnings: {report['blocking']} " \
                                   f"{report['degrading']}"


def test_a_corrupt_bundle_is_a_BLOCKING_no_go():
    """A partial upload must never present as a partial migration (D7)."""
    cfg, aw = _bundle()
    aw.write_json(BP.MANIFEST_JSON, {"files": [
        {"path": "export/vanished.json", "bytes": 10, "sha256": "deadbeef"}]})
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    assert report["verdict"] == NO_GO
    finding = _find(report, "bundle integrity")
    assert finding["grade"] == BLOCKING and not finding["ok"]
    assert "Re-copy" in finding["detail"], "the fix must be stated, not just the failure"


def test_a_live_import_without_a_state_schema_is_BLOCKING():
    """Without durable state a re-run cannot tell CREATE from UPDATE and may duplicate."""
    cfg, aw = _bundle(dry_run=False, state=False)
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    assert report["verdict"] == NO_GO
    assert _find(report, "migration state table")["grade"] == BLOCKING


def test_a_dry_run_without_a_state_schema_is_only_DEGRADING():
    """A first-look rehearsal must need no UC setup at all."""
    cfg, aw = _bundle(dry_run=True, state=False)
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "migration state table")
    assert finding["grade"] == DEGRADING
    assert report["verdict"] != NO_GO


def test_a_target_without_admin_rights_is_BLOCKING():
    class NoAdmin(RecordingClient):
        def get(self, path, params=None):
            if "scim/v2/Groups" in path:
                raise RuntimeError("PERMISSION_DENIED")
            return super().get(path, params)

    cfg, aw = _bundle()
    report = Preflight(NoAdmin(), cfg, aw, state=_store(cfg)).run()
    assert report["verdict"] == NO_GO
    assert _find(report, "workspace-admin")["grade"] == BLOCKING


def test_a_missing_account_GROUP_is_DEGRADING_and_NAMES_it():
    """An account GROUP is the only identity that can still be a prerequisite (Plan 6).

    Users and SPs are assigned automatically — the workspace SCIM POST creates them at the account,
    and an SP POST carrying `applicationId` adopts the existing account SP. So neither may appear
    here any more; only a group absent from the target account can hold up the migration, and it
    must NAME itself or the finding is unactionable.
    """
    cfg, aw = _bundle(files={
        BP.IDENTITY_CLASSIFICATION_JSON: {"identities": [
            {"identity_type": "group", "displayName": "present-grp", "kind": "account"},
            {"identity_type": "group", "displayName": "absent-grp", "kind": "account"},
            {"identity_type": "group", "displayName": "entra-grp", "kind": "account",
             "entra_backed": True},
            # Neither of these may ever be reported as a prerequisite again.
            {"identity_type": "user", "userName": "anyone@corp.com", "kind": "account"},
            {"identity_type": "service_principal", "applicationId": "umi-app", "kind": "account"},
        ]}})
    report = Preflight(_scim_client(account_groups=["present-grp"]), cfg, aw,
                       state=_store(cfg)).run()
    finding = _find(report, "account groups present")
    assert finding["grade"] == DEGRADING and not finding["ok"]
    assert report["verdict"] == GO_WITH_WARNINGS, "a missing account group must not BLOCK"
    affected = str(finding["affected_units"])
    assert "absent-grp" in affected and "entra-grp" in affected
    assert "present-grp" not in affected
    # each names WHO must fix it
    assert "Entra/SCIM" in affected and "account admin" in affected
    assert "anyone@corp.com" not in affected and "umi-app" not in affected, \
        "users and SPs are assigned automatically and must not be reported as prerequisites"


def test_an_account_group_shadowed_on_target_is_BLOCKING_in_preflight():
    """A workspace-local group squatting an account group's name blocks the assignment forever, so
    it must be caught BEFORE import writes anything."""
    cfg, aw = _bundle(files={
        BP.IDENTITY_CLASSIFICATION_JSON: {"identities": [
            {"identity_type": "group", "displayName": "corp-analysts", "kind": "account"}]}})
    client = _scim_client(groups=["corp-analysts"])
    report = Preflight(client, cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "not shadowed")
    assert finding["grade"] == BLOCKING and not finding["ok"]
    assert "corp-analysts" in str(finding["affected_units"])


def test_direct_mode_source_connectivity_is_BLOCKING_when_it_fails():
    """There'd be nothing to read — one of the four whole-run aborts."""
    class BadSource(RecordingClient):
        def get(self, path, params=None):
            raise RuntimeError("HTTP 401 from /oidc/v1/token")

    cfg, aw = _bundle(connectivity_mode="direct",
                      source={"workspace_url": "https://src", "client_id": "cid",
                              "spn_secret_value": "s"})
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg), source_client=BadSource()).run()
    assert report["verdict"] == NO_GO
    finding = _find(report, "source connectivity")
    assert finding["grade"] == BLOCKING
    assert "not mint" in finding["detail"] or "workspace admin" in finding["detail"]


def test_airgap_mode_never_checks_source_connectivity():
    cfg, aw = _bundle()
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "source connectivity")
    assert finding["ok"] and "never called" in finding["detail"]


def test_uc_references_are_DEGRADING_and_named():
    """The top cause of a 'successful' import producing a broken dashboard."""
    cfg, aw = _bundle(files={
        "export/dashboards/lakeview.json": {"units": [
            {"natural_key": "sales", "payload": {"serialized_dashboard": '{"datasets":[]}'}}]},
        "export/genie/spaces.json": {"units": [
            {"natural_key": "analytics", "payload": {"serialized_space": '{"tables":[]}'}}]},
        "export/dlt/pipelines.json": {"units": [
            {"natural_key": "bronze", "payload": {"catalog": "prod"}}]},
    })
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "Unity Catalog")
    assert finding["grade"] == DEGRADING and not finding["ok"]
    assert len(finding["affected_units"]) == 3
    assert "most common cause" in finding["detail"]


def test_akv_scopes_report_BOTH_prerequisites_separately():
    """The AAD-token cause and the vault-permission cause are fixed differently."""
    cfg, aw = _bundle(files={
        "export/secrets/scopes.json": {"units": [
            {"natural_key": "kv-scope", "payload": {
                "backend_type": "AZURE_KEYVAULT",
                "keyvault_metadata": {"dns_name": "https://v1.vault.azure.net/",
                                      "resource_id": "/subscriptions/x/v1"}}}]}})
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "Azure Key Vault")
    assert finding["grade"] == DEGRADING
    assert "AZURE AD token" in finding["detail"]
    assert "get`+`list` ON THE VAULT" in finding["detail"]
    assert "cross-region" in finding["detail"], "the region-1-vault reality must be stated"


def test_unresolvable_job_notebook_paths_are_DEGRADING_and_named():
    """The Jobs API accepts a bad path and fails at FIRST RUN, so this static check is the only
    thing that catches it before production (D14)."""
    cfg, aw = _bundle(files={
        "export/jobs.json": {"units": [
            {"natural_key": "etl", "payload": {"tasks": [
                {"task_key": "t", "notebook_task": {"notebook_path": "/Repos/x/gitfolder/nb"}}]}}]},
        "export/workspace/objects.json": {"units": []},
    })
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "notebook paths resolvable")
    assert finding["grade"] == DEGRADING and not finding["ok"]
    assert "etl → /Repos/x/gitfolder/nb" in finding["affected_units"]
    assert "FAIL AT FIRST RUN" in finding["detail"]


def test_repos_are_DEGRADING_and_legacy_dashboards_only_COSMETIC():
    """Grading matters: a repo can break a job, a legacy dashboard affects nothing else."""
    cfg, aw = _bundle(files={
        "export/workspace/repos.json": {"units": [{"natural_key": "/Repos/me/app"}]},
        "export/sql/legacy_dashboards.json": {"units": [{"natural_key": "old-dash"}]},
    })
    report = Preflight(_scim_client(), cfg, aw, state=_store(cfg)).run()
    assert _find(report, "Git repos")["grade"] == DEGRADING
    assert _find(report, "legacy SQL dashboards")["grade"] == "COSMETIC"


def test_a_target_with_no_warehouse_is_DEGRADING_not_blocking():
    cfg, aw = _bundle()
    report = Preflight(_scim_client(warehouses=()), cfg, aw, state=_store(cfg)).run()
    finding = _find(report, "warehouse availability")
    assert finding["grade"] == DEGRADING and not finding["ok"]
    assert report["verdict"] != NO_GO


def test_a_check_that_itself_errors_does_not_hide_the_others():
    """One broken check must not blind the operator to every other finding."""
    cfg, aw = _bundle()
    pf = Preflight(_scim_client(), cfg, aw, state=_store(cfg))
    pf.check_warehouses = lambda: (_ for _ in ()).throw(RuntimeError("boom"))
    report = pf.run()
    broken = _find(report, "warehouse availability")
    # The finding must be labelled with the CHECK's name, not "<lambda>" — otherwise the operator
    # cannot tell which check broke.
    assert not broken["ok"] and "failed to run" in broken["detail"]
    assert len(report["findings"]) > 5, "the other checks must still have run"
    assert report["verdict"] != NO_GO, "a broken check must not fabricate a blocker"


def test_preflight_writes_the_json_report_and_creates_nothing():
    cfg, aw = _bundle()
    client = _scim_client()
    Preflight(client, cfg, aw, state=_store(cfg)).run()
    assert os.path.isfile(os.path.join(aw.root, BP.PREFLIGHT_REPORT_JSON))
    # PLAN 7 §B2: no HTML twin any more — the JSON is the only preflight artifact.
    assert not os.path.isfile(os.path.join(aw.root, "preflight_report.html"))
    mutating = [c for c in client.calls if c[0] in ("POST", "PUT", "PATCH", "DELETE")]
    assert mutating == [], f"preflight must create NOTHING, but called: {mutating}"


# ── the import report set (§1a, D16) ───────────────────────────────────────

def _result_with_rows(rows):
    res = ImportResult("compute")
    for row in rows:
        res.add(row)
    return res


def _row(asset_type, key, status, **over):
    r = {"asset_type": asset_type, "natural_key": key, "family": "compute", "source_id": "s",
         "target_id": "t", "import_status": status, "action_taken": status,
         "fingerprint": "sha256:x", "note": "a note", "failure_category": "", "dry_run": False}
    r.update(over)
    return r


def test_import_results_json_is_joinable_on_asset_type_and_natural_key():
    """This is the ONLY thing Plan 4 needs from import — so the shape is asserted here, not assumed
    later."""
    cfg, aw = _bundle()
    results = [_result_with_rows([
        _row("cluster", "etl", "created"),
        _row("job", "nightly", "failed", failure_category="api_error"),
    ])]
    summary = {"run_id": "r1", "source_workspace_id": "111", "connectivity_mode": "airgap",
               "dry_run": False, "run_status": "completed", "elapsed_sec": 1.0,
               "totals": {"total": 2}, "per_phase": [r.as_dict() for r in results]}
    write_import_reports(aw, cfg, summary, results, {})

    payload = aw.read_json(BP.IMPORT_RESULTS_JSON)
    keys = {(u["asset_type"], u["natural_key"]) for u in payload["units"]}
    assert keys == {("cluster", "etl"), ("job", "nightly")}
    for unit in payload["units"]:
        for field in ("import_status", "action_taken", "target_id", "note"):
            assert field in unit, f"{field} missing — the report would have blank cells"
    assert payload["counts_by_status"] == {"created": 1, "failed": 1}
    assert payload["counts_by_asset_type"]["cluster"]["created"] == 1


def test_failures_sort_to_the_top_of_the_report():
    cfg, aw = _bundle()
    results = [_result_with_rows([
        _row("cluster", "aaa-alphabetically-first", "created"),
        _row("job", "zzz-last", "failed"),
    ])]
    write_import_reports(aw, cfg, {"run_id": "r", "totals": {}, "per_phase": []}, results, {})
    units = aw.read_json(BP.IMPORT_RESULTS_JSON)["units"]
    assert units[0]["import_status"] == "failed", "failures must lead, whatever they're called"


def test_every_artifact_is_written_including_the_workbook():
    """Import owns its OWN customer-readable output set, Excel included (D16). The slimmed set
    (PLAN 7 §B2) is json + xlsx + manual-actions — NO import_results.html any more."""
    cfg, aw = _bundle()
    results = [_result_with_rows([_row("cluster", "etl", "created"),
                                  _row("repo", "/Repos/me/app", "manual"),
                                  _row("acl", "clusters:etl", "skipped_no_object",
                                       failure_category="dab_redeploy")])]
    write_import_reports(aw, cfg, {"run_id": "r1", "source_workspace_id": "111",
                                   "dry_run": False, "run_status": "completed",
                                   "totals": {}, "per_phase": []}, results, {})
    for rel in (BP.IMPORT_RESULTS_JSON, BP.IMPORT_STATUS_XLSX, BP.MANUAL_ACTIONS_IMPORT_MD):
        p = os.path.join(aw.root, rel)
        assert os.path.isfile(p) and os.path.getsize(p) > 0, f"{rel} was not written"
    # the removed HTML must NOT reappear
    assert not os.path.isfile(os.path.join(aw.root, "import_results.html"))


def test_import_workbook_has_one_sheet_per_asset_type_named_like_inventory():
    """IMP-1: the import workbook must lay out ONE SHEET PER ASSET TYPE, named + ordered exactly
    like inventory.xlsx / export_status.xlsx — not one sheet per family (which lumped pools +
    policies + clusters into a single `compute` tab and diverged from the other two stages)."""
    from openpyxl import load_workbook
    from src.reports.inventory_view import _LABELS
    cfg, aw = _bundle()
    results = [_result_with_rows([
        _row("cluster", "etl", "created"),
        _row("instance_pool", "pool1", "created"),
        _row("cluster_policy", "pol1", "created"),
        _row("job", "nightly", "created"),
        _row("user", "a@x.com", "created"),
        _row("group", "grp", "created_with_warning"),
        _row("secret_scope", "kv", "created"),
        _row("genie_space", "space", "failed", failure_category="api_error"),
    ])]
    write_import_reports(aw, cfg, {"run_id": "r1", "source_workspace_id": "111",
                                   "dry_run": False, "run_status": "completed",
                                   "totals": {}, "per_phase": []}, results, {})
    wb = load_workbook(os.path.join(aw.root, BP.IMPORT_STATUS_XLSX))
    names = set(wb.sheetnames)
    # per-asset-type tabs, using the SAME labels inventory uses — NOT a single "compute" family tab
    assert "compute" not in names, "import workbook still groups by family"
    assert {_LABELS["clusters"], _LABELS["instance_pools"], _LABELS["cluster_policies"],
            _LABELS["jobs"], _LABELS["users"], _LABELS["groups"],
            _LABELS["secret_scopes"], _LABELS["genie_spaces"]} <= names
    # Instance Pools tab holds exactly its own rows, separate from clusters
    ws = wb[_LABELS["instance_pools"]]
    body = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    assert all(r[0] == "instance_pool" for r in body) and len(body) == 1


def test_the_manual_runbook_separates_the_four_kinds_of_outstanding_work():
    """Each needs a DIFFERENT action, so lumping them together would make the runbook useless."""
    cfg, aw = _bundle()
    results = [_result_with_rows([
        _row("repo", "/Repos/me/app", "manual", note="recreate by hand"),
        _row("job", "broken", "failed", note="quota exceeded"),
        _row("job", "degraded", "created_with_warning", note="notebook_path missing"),
        _row("acl", "directories:/Shared/.bundle/b", "skipped_no_object",
             failure_category="dab_redeploy", note="bundle-owned"),
    ])]
    write_import_reports(aw, cfg, {"run_id": "r1", "totals": {}, "per_phase": []}, results, {})
    md = open(os.path.join(aw.root, BP.MANUAL_ACTIONS_IMPORT_MD)).read()
    assert "Manual recreate" in md
    assert "retry_mode=failed_only" in md, "the retry instruction must be in the runbook"
    assert "Created but degraded" in md
    assert "import_assets=acls" in md and "retry_mode=skipped_only" in md


def test_a_dry_run_report_says_so_loudly_and_uses_its_own_filename():
    """Nobody should mistake a rehearsal for a migration, and a rehearsal must NOT overwrite a live
    run's report — it writes `import_status_dry_run.xlsx` (A1) and the runbook says DRY RUN."""
    cfg, aw = _bundle()
    results = [_result_with_rows([_row("cluster", "etl", "created", dry_run=True)])]
    write_import_reports(aw, cfg, {"run_id": "r1", "dry_run": True, "run_status": "completed",
                                   "totals": {}, "per_phase": []}, results, {})
    # the dry-run workbook is written; the LIVE name is left untouched
    assert os.path.isfile(os.path.join(aw.root, BP.IMPORT_STATUS_DRYRUN_XLSX))
    assert not os.path.isfile(os.path.join(aw.root, BP.IMPORT_STATUS_XLSX))
    md = open(os.path.join(aw.root, BP.MANUAL_ACTIONS_IMPORT_MD)).read()
    assert "DRY RUN" in md


def test_a_live_run_does_not_overwrite_the_dry_run_report():
    """A1 the other way round: a live run writes the plain name and leaves the rehearsal's alone."""
    cfg, aw = _bundle(dry_run=False)
    # rehearsal first
    write_import_reports(aw, cfg, {"run_id": "r1", "dry_run": True, "run_status": "completed",
                                   "totals": {}, "per_phase": []},
                         [_result_with_rows([_row("cluster", "etl", "created", dry_run=True)])], {})
    # then a live run
    write_import_reports(aw, cfg, {"run_id": "r1", "dry_run": False, "run_status": "completed",
                                   "totals": {}, "per_phase": []},
                         [_result_with_rows([_row("cluster", "etl", "created")])], {})
    assert os.path.isfile(os.path.join(aw.root, BP.IMPORT_STATUS_DRYRUN_XLSX)), \
        "the rehearsal's report must survive the live run"
    assert os.path.isfile(os.path.join(aw.root, BP.IMPORT_STATUS_XLSX))


def test_a_live_retry_run_writes_a_timestamped_set_and_preserves_the_full_run_report():
    """A live retry_mode!=off run must NOT clobber the full-run report — it writes its whole set
    tagged `_retry_<ts>` and leaves import_status.xlsx / import_results.json /
    manual_actions_import.md (the last full run) untouched."""
    cfg, aw = _bundle(dry_run=False)
    # a full live run first (canonical names)
    written_full = write_import_reports(
        aw, cfg, {"run_id": "r1", "source_workspace_id": "111", "dry_run": False,
                  "run_status": "completed", "totals": {}, "per_phase": []},
        [_result_with_rows([_row("cluster", "etl", "created")])], {})
    assert written_full["xlsx"] == BP.IMPORT_STATUS_XLSX
    assert written_full["json"] == BP.IMPORT_RESULTS_JSON

    # now a live RETRY run
    cfg.imports.retry_mode = "failed_only"
    written_retry = write_import_reports(
        aw, cfg, {"run_id": "r1", "source_workspace_id": "111", "dry_run": False,
                  "run_status": "completed", "totals": {}, "per_phase": []},
        [_result_with_rows([_row("dlt_pipeline", "bronze", "created")])], {})
    # the whole set is tagged, under the right dirs, and distinct from the canonical names
    assert written_retry["xlsx"].startswith("reports/import_status_retry_") \
        and written_retry["xlsx"].endswith(".xlsx")
    assert written_retry["json"].startswith("misc/import_results_retry_")
    assert written_retry["manual_actions"].startswith("reports/manual_actions_import_retry_")
    # the full-run files still exist AND the retry files exist alongside them
    for rel in (BP.IMPORT_STATUS_XLSX, BP.IMPORT_RESULTS_JSON, BP.MANUAL_ACTIONS_IMPORT_MD,
                written_retry["xlsx"], written_retry["json"], written_retry["manual_actions"]):
        assert os.path.isfile(os.path.join(aw.root, rel)), f"{rel} missing"
    # the preserved full-run json still shows the ORIGINAL outcome, not the retry's
    assert aw.read_json(BP.IMPORT_RESULTS_JSON)["units"][0]["asset_type"] == "cluster"
    assert aw.read_json(written_retry["json"])["units"][0]["asset_type"] == "dlt_pipeline"


def test_a_retry_report_is_scoped_to_the_units_it_attempted():
    """The retry file must NOT re-print the whole inventory as 'not outstanding' — only the units
    the retry actually attempted appear, and the roll-up counts match."""
    cfg, aw = _bundle(dry_run=False)
    cfg.imports.retry_mode = "failed_only"
    rows = [
        _row("dlt_pipeline", "bronze", "created"),                       # attempted → shown
        _row("cluster", "etl", "skipped", retry_out_of_scope=True,       # not in bucket → hidden
             action_taken="Not in this retry's work list", note="not outstanding"),
    ]
    written = write_import_reports(
        aw, cfg, {"run_id": "r1", "source_workspace_id": "111", "dry_run": False,
                  "run_status": "completed", "totals": {"total": 2, "skipped": 1, "created": 1},
                  "per_phase": []},
        [_result_with_rows(rows)], {})
    payload = aw.read_json(written["json"])
    keys = {(u["asset_type"], u["natural_key"]) for u in payload["units"]}
    assert keys == {("dlt_pipeline", "bronze")}, "the retry report must show ONLY attempted units"
    assert payload["totals"]["total"] == 1 and payload["totals"]["created"] == 1
    assert payload["scoped_to_retry"] is True
    # the workbook likewise has no out-of-scope cluster row
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(aw.root, written["xlsx"]))
    seen = set()
    for name in wb.sheetnames:
        for r in wb[name].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                seen.add(str(r[0]))
    assert "cluster" not in seen, "an out-of-scope unit leaked into the retry workbook"


def test_a_dry_run_is_not_treated_as_a_retry_even_with_retry_mode_set():
    """retry tagging is for LIVE runs; a dry rehearsal keeps its own _dry_run name regardless."""
    cfg, aw = _bundle(dry_run=True)
    cfg.imports.retry_mode = "failed_only"
    written = write_import_reports(
        aw, cfg, {"run_id": "r1", "dry_run": True, "run_status": "completed",
                  "totals": {}, "per_phase": []},
        [_result_with_rows([_row("cluster", "etl", "created", dry_run=True)])], {})
    assert written["xlsx"] == BP.IMPORT_STATUS_DRYRUN_XLSX
    assert "retry_" not in written["json"] and "retry_" not in written["xlsx"]


def test_an_aborted_run_is_visibly_partial():
    """An abort must never look clean — the results json is marked aborted."""
    cfg, aw = _bundle()
    results = [_result_with_rows([_row("cluster", "etl", "created")])]
    write_import_reports(aw, cfg, {"run_id": "r1", "run_status": "aborted",
                                   "abort_reason": "driver killed", "totals": {},
                                   "per_phase": []}, results, {})
    payload = aw.read_json(BP.IMPORT_RESULTS_JSON)
    assert payload["run_status"] == "aborted"


def test_acl_parity_is_folded_into_the_workbook_as_a_sheet():
    """D-1: parity is no longer a standalone acl_parity_report.{json,html}; it becomes the
    'ACL Parity' sheet of import_status.xlsx and the standalone files are NOT written."""
    from openpyxl import load_workbook
    cfg, aw = _bundle(dry_run=False)
    results = [_result_with_rows([_row("acl", "clusters:etl", "created")])]
    parity = {"run_id": "r1", "objects_checked": 1, "counts": {"match": 1},
              "known_limitation": "a platform quirk",
              "objects": [{"perm_object_type": "clusters", "object": "etl", "verdict": "match",
                           "missing_on_target": [], "extra_on_target": []}]}
    write_import_reports(aw, cfg, {"run_id": "r1", "source_workspace_id": "111", "dry_run": False,
                                   "run_status": "completed", "totals": {}, "per_phase": []},
                         results, {"acl_parity": parity})
    wb = load_workbook(os.path.join(aw.root, BP.IMPORT_STATUS_XLSX))
    assert "ACL Parity" in wb.sheetnames
    assert not os.path.isfile(os.path.join(aw.root, "acl_parity_report.json"))
    assert not os.path.isfile(os.path.join(aw.root, "acl_parity_report.html"))


def test_deleted_in_source_is_reported_as_review_not_deletion():
    cfg, aw = _bundle()
    results = [_result_with_rows([_row("cluster", "etl", "created")])]
    write_import_reports(aw, cfg, {"run_id": "r1", "totals": {}, "per_phase": []}, results,
                         {"deleted_in_source": {"job": ["retired-job"]}})
    md = open(os.path.join(aw.root, BP.MANUAL_ACTIONS_IMPORT_MD)).read()
    assert "Deleted in source" in md and "retired-job" in md
    assert "allow_deletes=true" in md, "the report must say deletion did NOT happen"
